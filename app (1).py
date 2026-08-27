import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
import matplotlib.gridspec as gridspec
import matplotlib.patheffects as pe
from pycirclize import Circos
import warnings
warnings.filterwarnings("ignore")

from scipy import stats
from scipy.optimize import minimize_scalar
from scipy.cluster.hierarchy import dendrogram, linkage
from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression
from sklearn.manifold import MDS
from statsmodels.stats.multitest import multipletests
import io, re, os, zipfile, tempfile
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────── DEFAULT COLOR PALETTES ──────────────────────────────────
BG_DARK     = "#0D1117"
PANEL_DARK  = "#161B22"
PANEL_MID   = "#21262D"
ACCENT_TEAL = "#2DD4BF"
ACCENT_GOLD = "#F59E0B"
ACCENT_CORAL= "#F87171"
ACCENT_BLUE = "#60A5FA"
ACCENT_LIME = "#84CC16"
ACCENT_PURP = "#A78BFA"
ACCENT_PINK = "#F472B6"
ACCENT_CYAN = "#22D3EE"
ACCENT_ORG  = "#FB923C"
TEXT_MAIN   = "#E6EDF3"
TEXT_DIM    = "#8B949E"
GRID_LINE   = "#30363D"
SIG_COLOR   = "#F59E0B"

# Light-theme equivalents -- used for ALL plots (PCA must NOT default to dark)
BG_LIGHT      = "#FFFFFF"
PANEL_LIGHT   = "#FAFBFF"
TEXT_DARK     = "#2D3142"
GRID_LIGHT    = "#E8ECF4"
ACCENT_TEAL_D = "#0F766E"
ACCENT_GOLD_D = "#B45309"
ACCENT_PINK_D = "#BE185D"

MODEL_COLORS = {
    "OLS":        "#60A5FA",
    "MLM":        "#2DD4BF",
    "EMMAX":      "#F472B6",
    "FarmCPU":    "#84CC16",
    "GEMMA":      "#E879F9",
    "BLINK":      "#FB923C",
    "mrMLM":      "#A78BFA",
    "FASTmrMLM":  "#F59E0B",
}

CHR_PALETTE = [
    "#4E79A7","#F28E2B","#E15759","#76B7B2","#59A14F",
    "#EDC948","#B07AA1","#FF9DA7","#9C755F","#BAB0AC",
    "#499894","#86BCB6","#D4A6C8","#FFBE7D","#8CD17D",
]

ALL_MODELS = ["OLS", "MLM", "EMMAX", "FarmCPU", "GEMMA", "BLINK", "mrMLM", "FASTmrMLM"]

def style_fig_white(fig, font_size=9, font_color="#2D3142", grid_color="#E8ECF4"):
    fig.patch.set_facecolor("#FFFFFF")
    for ax in fig.get_axes():
        ax.set_facecolor("#FAFBFF")
        ax.tick_params(colors=font_color, labelsize=font_size)
        ax.xaxis.label.set_color(font_color)
        ax.yaxis.label.set_color(font_color)
        ax.title.set_color(font_color)
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
        ax.grid(color=grid_color, linewidth=0.5, linestyle="--", alpha=0.7)
    return fig

def apply_custom_style(fig, cfg):
    bg = cfg.get("bg_color", "#FFFFFF")
    panel = cfg.get("panel_color", "#FAFBFF")
    fc = cfg.get("font_color", "#2D3142")
    fs = cfg.get("font_size", 9)
    gc = cfg.get("grid_color", "#E8ECF4")
    fig.patch.set_facecolor(bg)
    for ax in fig.get_axes():
        ax.set_facecolor(panel)
        ax.tick_params(colors=fc, labelsize=fs)
        ax.xaxis.label.set_color(fc)
        ax.yaxis.label.set_color(fc)
        ax.title.set_color(fc)
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
        ax.grid(color=gc, linewidth=0.5, linestyle="--", alpha=0.7)
        ax.xaxis.label.set_fontsize(fs)
        ax.yaxis.label.set_fontsize(fs)
        ax.title.set_fontsize(fs + 2)
    return fig

# ─────────────────────── DATA PARSING ────────────────────────────────────────────
def parse_data(file_obj=None, text_input=None):
    raw = None
    if file_obj is not None:
        try:
            raw = pd.read_csv(file_obj.name, sep=None, engine="python")
        except Exception:
            raw = pd.read_csv(file_obj.name, sep="\t")
    elif text_input and text_input.strip():
        try:
            raw = pd.read_csv(StringIO(text_input), sep=None, engine="python")
        except Exception:
            raw = pd.read_csv(StringIO(text_input), sep="\t")
    else:
        raise ValueError("No data provided.")

    chrom_info = {}
    if len(raw) > 1:
        second_row = raw.iloc[1].astype(str)
        chrom_vals = second_row.iloc[2:].astype(str)
        if len(chrom_vals) > 0 and chrom_vals.str.match(r'^\d+$').all():
            for i, col in enumerate(raw.columns[2:]):
                try:
                    chrom_info[col] = int(second_row.iloc[2 + i])
                except:
                    chrom_info[col] = 1
            raw = raw.drop(index=1).reset_index(drop=True)

    cols = list(raw.columns)
    id_col = cols[0]
    phen_col = cols[1]
    snp_cols = cols[2:]

    raw = raw[pd.to_numeric(raw[phen_col], errors="coerce").notna()].copy()
    raw[phen_col] = raw[phen_col].astype(float)
    geno = raw[snp_cols].apply(pd.to_numeric, errors="coerce")

    # Only these are unambiguous missing-value sentinels. -1 is deliberately
    # NOT in this list: many marker matrices (rrBLUP/GAPIT-style) use -1/0/1
    # as REAL genotype calls (homozygous ref / het / homozygous alt), not a
    # missing code. Blindly nuking -1 as missing in that case wipes out
    # ~30-55% of real genotype calls and corrupts MAF, kinship, and PCA
    # downstream (all of which assume a clean 0/1/2 dosage scale).
    geno = geno.where(~geno.isin([-9, -99, -999]), np.nan)

    # Detect the coding scheme from the observed value range and normalize
    # everything to a 0/1/2 additive dosage scale so downstream math is
    # always consistent regardless of which format the input used.
    obs_vals = geno.values[~pd.isna(geno.values)]
    if obs_vals.size > 0 and obs_vals.min() >= -1 and obs_vals.max() <= 1:
        # -1/0/1 coding detected -> rescale to 0/1/2
        geno = geno + 1
    else:
        # 0/1/2 coding: here -1 really is a missing-value sentinel
        geno = geno.where(geno != -1, np.nan)

    df = pd.DataFrame()
    df["ID"] = raw[id_col].values
    df["Phenotype"] = raw[phen_col].values
    for c in snp_cols:
        df[c] = geno[c].values

    chroms, positions = [], []
    # Fallback position counters MUST be per-chromosome, not a single running
    # index over the whole marker file. A global counter (the old
    # `len(positions) + 1`) makes later chromosomes carry huge positional
    # offsets (e.g. chr9 markers numbered ~3100-3400 instead of 1-300), which
    # plot_manhattan() masks by subtracting each chromosome's own min before
    # plotting, but plot_circos_density() does not -- producing a genuinely
    # empty arc at the start of every sector sized to the wrong (inflated)
    # chromosome length. Keeping the counter per-chromosome fixes it at the
    # source instead of relying on every downstream plot to compensate.
    chr_counters = {}

    def _next_idx(c):
        chr_counters[c] = chr_counters.get(c, 0) + 1
        return chr_counters[c]

    for m in snp_cols:
        if m in chrom_info:
            c = chrom_info[m]
            chroms.append(c)
            positions.append(_next_idx(c))
        else:
            m_clean = str(m).strip()
            mat = re.match(r"[Cc]hr(\d+)[_\-](\d+)", m_clean)
            if mat:
                c = int(mat.group(1))
                chroms.append(c)
                positions.append(int(mat.group(2)))
            else:
                num = re.findall(r"\d+", m_clean)
                c = int(num[0]) if num else 1
                chroms.append(c)
                positions.append(int(num[1]) if len(num) > 1 else _next_idx(c))

    return df, snp_cols, np.array(chroms), np.array(positions)

# ─────────────────────── UTILITY ─────────────────────────────────────────────────
def impute_genotypes(geno_matrix):
    mat = geno_matrix.copy().astype(float)
    for j in range(mat.shape[1]):
        col = mat[:, j]
        m = np.nanmean(col)
        col[np.isnan(col)] = m if not np.isnan(m) else 0
        mat[:, j] = col
    return mat

def compute_maf(G):
    freqs = np.nanmean(G, axis=0) / 2
    return np.where(freqs > 0.5, 1 - freqs, freqs)

def compute_call_rate(G):
    return 1 - np.isnan(G).mean(axis=0)

def build_kinship(G):
    p = np.nanmean(G, axis=0) / 2
    Z = G - 2 * p
    denom = 2 * np.sum(p * (1 - p))
    if denom == 0:
        denom = 1
    return Z @ Z.T / denom

def compute_pve_all(y, G, K=None, X0=None):
    """
    PVE (%) for every marker via genuine variance-component partitioning
    instead of a raw squared Pearson correlation.

    A plain r^2 between one marker and y only equals a valid PVE estimate
    when the model has no other covariates and no relatedness structure
    (it happens to coincide with SS_marker/SS_total for a bare
    single-predictor OLS fit). Once a kinship random effect or Q
    covariates are in the model, that raw correlation double-counts
    variance already absorbed by relatedness/structure and no longer
    matches what GAPIT/GEMMA-style PVE figures report.

    This instead computes, for each marker j:
        PVE_j = (RSS_null - RSS_full_j) / SS_total(y) * 100
    where RSS_null is the residual sum of squares of the model with just
    the intercept (+ X0 covariates, if given), RSS_full_j additionally
    includes marker j, and SS_total(y) is the total phenotypic sum of
    squares -- the same denominator convention GREML/GCTA-style "variance
    explained" figures use, so PVE values from different markers are
    directly comparable and summable, unlike per-marker r^2.

    If K (kinship) is supplied, y/X0/G are first weighted by the
    REML-estimated genetic/residual variance ratio (delta), exactly as
    mixed_model_gwas does, so a marker's PVE is *conditional on* the
    polygenic background already explained by relatedness rather than
    re-claiming variance the random effect already accounts for.
    """
    y = np.asarray(y, dtype=float)
    n, m = G.shape
    if X0 is None:
        X0 = np.ones((n, 1))

    SS_total = float(np.sum((y - y.mean()) ** 2))
    if SS_total <= 0:
        return np.zeros(m)

    if K is not None:
        try:
            evals, evecs = eigh_kinship(K)
            _, _, delta, _, _ = emma_reml(y, K, X0, evals=evals, evecs=evecs)
            w = 1.0 / (evals + delta)
            sw = np.sqrt(w)
            yt = (evecs.T @ y) * sw
            X0t = (evecs.T @ X0) * sw[:, None]
            Gt = (evecs.T @ G) * sw[:, None]
        except Exception:
            yt, X0t, Gt = y, X0, G
    else:
        yt, X0t, Gt = y, X0, G

    beta0, *_ = np.linalg.lstsq(X0t, yt, rcond=None)
    rss0 = float(np.sum((yt - X0t @ beta0) ** 2))

    pve = np.zeros(m)
    for j in range(m):
        xj = Gt[:, j]
        if np.std(xj) < 1e-8:
            continue
        Xf = np.column_stack([X0t, xj])
        try:
            beta, *_ = np.linalg.lstsq(Xf, yt, rcond=None)
        except np.linalg.LinAlgError:
            continue
        rss = float(np.sum((yt - Xf @ beta) ** 2))
        pve[j] = max(0.0, rss0 - rss) / SS_total * 100
    return pve


RENDER_DPI = 300  # publication-quality; was 150
EXPORT_DPI = 400  # all downloadable PNG/figure files must exceed 300 dpi

def fig_to_png_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=RENDER_DPI, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    return buf.read()

def fig_to_pil(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=RENDER_DPI, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    from PIL import Image
    return Image.open(buf).copy()

def get_top_mta(pvals, n=15):
    return np.argsort(pvals)[:min(n, len(pvals))]

def get_significant_idx(pvals, lod_threshold=None, fallback_top_n=10,
                         pve_vals=None, pve_threshold=None):
    """
    Indices of markers that actually PASS the significance threshold --
    this is what should be highlighted as MTAs, not a blind top-N slice
    that highlights N markers regardless of whether any of them are real.

    lod_threshold is on the LOD scale (see neglog10p_to_lod) and is meant
    to be the permutation-derived empirical threshold -- the primary
    significance criterion for this pipeline, the same role a LOD
    threshold plays in linkage/QTL mapping. Returns indices sorted by
    increasing p-value (most significant first). If no permutation
    threshold is available (e.g. it failed to compute), falls back to a
    small top-N so plots don't render silently empty -- this is a safety
    net, not the intended mode of operation.

    pve_vals / pve_threshold add a SECOND, independent filter: if both are
    given, a marker must ALSO explain at least pve_threshold % of
    phenotypic variance to be kept as an MTA, on top of passing the LOD
    threshold. pve_threshold of None/0 disables this filter.
    """
    pvals = np.asarray(pvals)
    if lod_threshold is None:
        idx = get_top_mta(pvals, fallback_top_n)
    else:
        log_p = -np.log10(np.clip(pvals, 1e-300, 1))
        lod_vals = neglog10p_to_lod(log_p)
        idx = np.where(lod_vals >= lod_threshold)[0]
        idx = idx[np.argsort(pvals[idx])]

    if pve_vals is not None and pve_threshold:
        pve_vals = np.asarray(pve_vals)
        idx = idx[pve_vals[idx] >= pve_threshold]

    return idx

def compute_thresholds(n_snps, model_name=""):
    """
    Compute and explain Bonferroni and suggestive thresholds.
    Returns thresholds and explanation text.
    """
    bonferroni = 0.05 / n_snps
    suggestive = 1.0 / n_snps  # 1 expected false positive genome-wide

    model_notes = {
        "OLS":     "No structure/relatedness correction — baseline; expect inflation (λ > 1).",
        "MLM":     "Q (kinship PCs, fixed) + K (polygenic random effect, REML) — Yu et al. 2006.",
        "EMMAX":   "Pure kinship P3D mixed model; variance components estimated once genome-wide.",
        "FarmCPU": "Iterative FEM/REM with bin-based pseudo-QTN cofactors (Liu et al. 2016).",
        "GEMMA":   "Exact per-marker REML LMM (Zhou & Stephens 2012); no post-hoc GC rescaling.",
    }
    note = model_notes.get(model_name, "")

    explanation = (
        f"──────────────────────────────────────────────\n"
        f"  Threshold Calculations — {model_name}\n"
        f"──────────────────────────────────────────────\n"
        f"  SNPs tested (m) : {n_snps:,}\n\n"
        f"  📌 Bonferroni threshold:\n"
        f"     α / m = 0.05 / {n_snps:,} = {bonferroni:.3e}\n"
        f"     -log₁₀(p) = {-np.log10(bonferroni):.2f}\n"
        f"     Controls genome-wide Type-I error at 5%.\n"
        f"     Assumes all tests are independent.\n\n"
        f"  📌 Suggestive threshold:\n"
        f"     1 / m = 1 / {n_snps:,} = {suggestive:.3e}\n"
        f"     -log₁₀(p) = {-np.log10(suggestive):.2f}\n"
        f"     Expects ~1 false positive per genome scan.\n"
        f"     Lander & Kruglyak (1995) convention.\n\n"
        f"  ℹ️  Model note: {note}\n"
        f"──────────────────────────────────────────────\n"
    )
    return bonferroni, suggestive, explanation

# ─────────────────────── EMMA / REML MIXED-MODEL ENGINE ──────────────────────────
# Shared machinery for MLM / EMMAX / GEMMA, following the eigendecomposition
# approach of Kang et al. 2008 (EMMA) and its extensions (EMMAX, FaST-LMM,
# GEMMA). The kinship matrix is decomposed ONCE; every subsequent variance-
# component search reuses the eigenvalues, which is what makes these methods
# tractable at genome scale instead of refitting an n x n GLS from scratch
# per marker.

def eigh_kinship(K):
    """Symmetric eigendecomposition of the kinship matrix. Eigenvalues are
    floored at a small positive number to keep downstream log/inversion
    steps numerically stable (K can be near-singular for small panels)."""
    n = K.shape[0]
    Ks = (K + K.T) / 2.0
    evals, evecs = np.linalg.eigh(Ks)
    evals = np.clip(evals, 1e-8, None)
    return evals, evecs

def _profile_loglik(log_delta, evals, yt, Xt, reml=True):
    """REML (or ML) profile log-likelihood as a function of delta =
    sigma_e^2 / sigma_g^2, evaluated in the rotated (eigenvector) basis
    where the covariance structure becomes diagonal: var(yt_i) = D_i + delta.
    """
    delta = np.exp(log_delta)
    w = 1.0 / (evals + delta)
    sw = np.sqrt(w)
    Xw = Xt * sw[:, None]
    yw = yt * sw
    beta, *_ = np.linalg.lstsq(Xw, yw, rcond=None)
    resid = yw - Xw @ beta
    n = len(yt)
    q = Xt.shape[1]
    df = max(n - q, 1) if reml else n
    rss = float(np.sum(resid ** 2))
    sigma2 = rss / df if df > 0 else rss / max(n, 1)
    if sigma2 <= 0:
        sigma2 = 1e-12
    ll = -0.5 * (df * np.log(2 * np.pi * sigma2) + np.sum(np.log(evals + delta)) + df)
    if reml:
        XtWX = Xw.T @ Xw
        sign, logdet = np.linalg.slogdet(XtWX)
        if sign > 0:
            ll -= 0.5 * logdet
    return ll

def emma_reml(y, K, X=None, evals=None, evecs=None):
    """
    Estimate the genetic/residual variance ratio delta via REML, using the
    eigendecomposition of K (Kang et al. 2008). Returns eigvals, eigvecs,
    delta_hat, sigma_g2_hat, sigma_e2_hat.
    """
    n = len(y)
    if X is None:
        X = np.ones((n, 1))
    if evals is None or evecs is None:
        evals, evecs = eigh_kinship(K)
    yt = evecs.T @ y
    Xt = evecs.T @ X

    res = minimize_scalar(lambda ld: -_profile_loglik(ld, evals, yt, Xt, reml=True),
                           bounds=(-10, 10), method="bounded",
                           options={"xatol": 1e-4})
    delta_hat = float(np.exp(res.x))

    w = 1.0 / (evals + delta_hat)
    sw = np.sqrt(w)
    Xw = Xt * sw[:, None]
    yw = yt * sw
    beta, *_ = np.linalg.lstsq(Xw, yw, rcond=None)
    resid = yw - Xw @ beta
    df = max(n - X.shape[1], 1)
    sigma_g2 = float(np.sum(resid ** 2) / df)
    sigma_e2 = sigma_g2 * delta_hat
    return evals, evecs, delta_hat, sigma_g2, sigma_e2

def _wald_test_rotated(y_col, X_cols, evals, delta):
    """GLS Wald test for the last column of X_cols (the marker), with y and
    X already rotated into the eigenvector basis. Weights 1/(D_i+delta) make
    this an ordinary weighted least squares problem — the whole reason the
    eigendecomposition trick works."""
    w = 1.0 / (evals + delta)
    sw = np.sqrt(w)
    Xw = X_cols * sw[:, None]
    yw = y_col * sw
    try:
        XtX_inv = np.linalg.inv(Xw.T @ Xw)
    except np.linalg.LinAlgError:
        return 1.0, 0.0, np.nan
    beta = XtX_inv @ (Xw.T @ yw)
    resid = yw - Xw @ beta
    dfree = max(len(y_col) - X_cols.shape[1], 1)
    sigma2 = float(np.sum(resid ** 2) / dfree)
    se = float(np.sqrt(max(sigma2 * XtX_inv[-1, -1], 0)))
    b = float(beta[-1])
    t_stat = b / (se + 1e-300)
    p = float(2 * stats.t.sf(abs(t_stat), df=dfree))
    return p, b, se

def mixed_model_gwas(y, G, K, covariates=None):
    """
    P3D mixed-model GWAS (Population Parameters Previously Determined):
    variance components are estimated ONCE from the null model (no marker
    effect), then reused for every marker's GLS test. This is what real
    MLM/EMMAX software does by default (TASSEL, GAPIT, EMMAX itself) since
    per-marker REML gives negligible accuracy gain for a large speed cost.
    """
    n, m = G.shape
    X0 = np.ones((n, 1))
    if covariates is not None and covariates.shape[1] > 0:
        X0 = np.column_stack([X0, covariates])

    evals, evecs, delta, sigma_g2, sigma_e2 = emma_reml(y, K, X0)
    yt = evecs.T @ y
    X0t = evecs.T @ X0
    Gt = evecs.T @ G  # rotate all markers once — O(n^2 m), not O(n^3) per marker

    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
    for j in range(m):
        if G[:, j].std() < 1e-8:
            continue
        Xcols = np.column_stack([X0t, Gt[:, j]])
        p, b, se = _wald_test_rotated(yt, Xcols, evals, delta)
        pvals[j] = p; betas[j] = b; ses[j] = se
    return pvals, betas, ses, delta

# ─────────────────────── PERMUTATION (LOD) THRESHOLD ──────────────────────────────
def permutation_threshold_lod(y, G, K, extra_covariates=None, n_perm=100, alpha=0.05, seed=42):
    """
    Empirical permutation significance threshold (Churchill & Doerge 1994
    style -- the same logic used for LOD thresholds in linkage/QTL mapping),
    expressed on the LOD scale via neglog10p_to_lod so it's directly
    comparable to a QTL-mapping LOD threshold. This is meant to be THE
    primary significance criterion for this pipeline -- which MTAs get
    highlighted should be driven by whether they pass this threshold, not
    by an arbitrary top-N slice.

    Phenotype is shuffled n_perm times; for each permutation the maximum
    -log10(p) across the whole genome is recorded via the fast P3D mixed-
    model engine (variance components estimated once on the REAL data and
    held fixed across permutations -- standard practice for tractable
    mixed-model permutation, since re-running REML per permutation per
    marker would be computationally prohibitive). The threshold is the
    (1-alpha) quantile of that null max-statistic distribution, converted
    to LOD units.
    """
    rng = np.random.default_rng(seed)
    n, m = G.shape
    X0 = np.ones((n, 1))
    if extra_covariates is not None and extra_covariates.shape[1] > 0:
        X0 = np.column_stack([X0, extra_covariates])

    evals, evecs, delta, _, _ = emma_reml(y, K, X0)
    X0t = evecs.T @ X0
    Gt = evecs.T @ G
    valid_markers = np.where(G.std(axis=0) > 1e-8)[0]

    max_lp = np.zeros(n_perm)
    for b in range(n_perm):
        y_perm = rng.permutation(y)
        yt = evecs.T @ y_perm
        best = 0.0
        for j in valid_markers:
            Xcols = np.column_stack([X0t, Gt[:, j]])
            p, _, _ = _wald_test_rotated(yt, Xcols, evals, delta)
            lp = -np.log10(max(p, 1e-300))
            if lp > best:
                best = lp
        max_lp[b] = best

    threshold_neglog10p = float(np.quantile(max_lp, 1 - alpha))
    threshold_lod = float(neglog10p_to_lod(threshold_neglog10p))
    return threshold_lod, threshold_neglog10p, max_lp


def ols_gwas(y, G):
    """Simple linear regression per marker, no correction for structure or
    relatedness. The baseline every other model is judged against."""
    n, m = G.shape
    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
    for j in range(m):
        x = G[:, j] - G[:, j].mean()
        if x.std() < 1e-8: continue
        slope, _, _, p, se = stats.linregress(x, y)
        pvals[j] = p; betas[j] = slope; ses[j] = se
    return pvals, betas, ses

def get_kinship_pcs(K, n_pca, n_samples):
    """Top n_pca eigenvectors of the kinship matrix, used as fixed-effect
    'Q' structure covariates by MLM, mrMLM, FASTmrMLM, and (optionally)
    GEMMA. Shared here so the per-model PCA covariate count set in the UI
    (Cofactors tab) is applied identically across every model that uses it."""
    n_pca = max(min(int(n_pca), n_samples - 4), 0)
    if n_pca <= 0:
        return None
    evals_K, evecs_K = eigh_kinship(K)
    order = np.argsort(evals_K)[::-1]
    return evecs_K[:, order[:n_pca]]

def mlm_gwas(y, G, K, n_pca=3):
    """Q + K mixed linear model (Yu et al. 2006): top kinship eigenvectors
    as fixed 'Q' structure covariates, PLUS the polygenic random effect
    (via REML-estimated variance components) rather than eigenvectors
    substituting for it. This is the actual unified MLM, not a PCA-only
    correction. n_pca sets how many kinship PCs are used as covariates."""
    n = len(y)
    Q = get_kinship_pcs(K, n_pca, n)
    pvals, betas, ses, _ = mixed_model_gwas(y, G, K, covariates=Q)
    return pvals, betas, ses

def emmax_gwas(y, G, K):
    """EMMAX (Kang et al. 2008): pure kinship-based P3D mixed model, no
    separate Q structure covariates — relatedness alone absorbs population
    structure. Same REML/eigendecomposition engine as MLM, without the
    fixed PCs."""
    pvals, betas, ses, _ = mixed_model_gwas(y, G, K, covariates=None)
    return pvals, betas, ses

def _rem_prune_qtns(y, G, K, qtn_idx, alpha=0.01):
    """
    REM (Random Effect Model) step of FarmCPU: re-test each candidate
    pseudo-QTN's own effect in a kinship-based mixed model, with the other
    current pseudo-QTNs held in as fixed covariates, and drop any QTN that
    is no longer significant once the polygenic random effect is in the
    model. Real FarmCPU alternates exactly this -- a REM re-optimization of
    the pseudo-QTN set -- with the FEM genome scan; without it (fixed
    effects only, as in the previous version of this function) pseudo-QTNs
    that are only "significant" because of relatedness/background genetic
    correlation, rather than a real local effect, never get removed.
    """
    if not qtn_idx:
        return qtn_idx
    n = len(y)
    keep = []
    for idx in qtn_idx:
        others = [q for q in qtn_idx if q != idx]
        cov = None
        if others:
            cov = G[:, others] - G[:, others].mean(axis=0)
        try:
            p, _, _, _ = mixed_model_gwas(y, G[:, [idx]], K, covariates=cov)
        except Exception:
            keep.append(idx)  # fit failed -- don't silently drop a candidate
            continue
        if p[0] <= alpha:
            keep.append(idx)
    return keep

def farmcpu_gwas(y, G, K, chroms=None, positions=None, max_iter=8, bin_size=1_000_000,
                  rem_alpha=0.01):
    """
    FarmCPU (Liu et al. 2016): iterates (1) a bin-based candidate scan --
    a Fixed Effect Model (FEM) test of every marker with the current
    pseudo-QTNs as covariates, followed by one candidate kept per bin --
    and (2) a Random Effect Model (REM) step that re-tests each candidate
    pseudo-QTN in a kinship-based mixed model and drops any that aren't
    significant once relatedness is accounted for. This is the actual
    FEM/REM alternation FarmCPU is defined by, not FEM alone: the REM step
    is what keeps pseudo-QTNs that are only correlated with y through
    background relatedness (rather than a true local effect) from being
    locked in as covariates for the rest of the run.
    """
    n, m = G.shape
    if positions is None:
        positions = np.arange(m)
    if chroms is None:
        chroms = np.zeros(m, dtype=int)

    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
    pseudo_qtns, prev_qtns = [], None

    for _ in range(max_iter):
        X = np.ones((n, 1))
        if pseudo_qtns:
            cov = G[:, pseudo_qtns] - G[:, pseudo_qtns].mean(axis=0)
            X = np.column_stack([X, cov])

        pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
        qtn_set = set(pseudo_qtns)
        for j in range(m):
            x = G[:, j]
            if x.std() < 1e-8:
                continue
            # A marker that is currently a pseudo-QTN is tested leave-one-out
            # (covariates = the OTHER pseudo-QTNs), not skipped -- otherwise
            # the true causal marker, which is exactly the one most likely
            # to have been selected as its own covariate, ends up at p=1.
            if j in qtn_set:
                others = [q for q in pseudo_qtns if q != j]
                if others:
                    cov = G[:, others] - G[:, others].mean(axis=0)
                    Xj = np.column_stack([np.ones((n, 1)), cov])
                else:
                    Xj = np.ones((n, 1))
            else:
                Xj = X
            Xf = np.column_stack([Xj, x])
            try:
                beta = np.linalg.lstsq(Xf, y, rcond=None)[0]
                XtX_inv = np.linalg.pinv(Xf.T @ Xf)
            except np.linalg.LinAlgError:
                continue
            resid = y - Xf @ beta
            dfree = max(n - Xf.shape[1], 1)
            sigma2 = float(np.sum(resid ** 2) / dfree)
            se = float(np.sqrt(max(sigma2 * XtX_inv[-1, -1], 0)))
            b = float(beta[-1])
            t_stat = b / (se + 1e-300)
            pvals[j] = float(2 * stats.t.sf(abs(t_stat), df=dfree))
            betas[j] = b; ses[j] = se

        bonf = 0.05 / m
        cand = np.where(pvals < bonf)[0]
        if len(cand) == 0:
            cand = np.argsort(pvals)[:5]

        bins = {}
        for idx in cand:
            key = (int(chroms[idx]), int(positions[idx] // bin_size))
            if key not in bins or pvals[idx] < pvals[bins[key]]:
                bins[key] = idx
        new_qtns = sorted(bins.values(), key=lambda i: pvals[i])[:10]

        # REM step: prune bin-selected candidates that don't hold up once
        # tested in a kinship random-effect model (see _rem_prune_qtns).
        new_qtns = _rem_prune_qtns(y, G, K, new_qtns, alpha=rem_alpha)

        if prev_qtns is not None and set(new_qtns) == set(prev_qtns):
            pseudo_qtns = new_qtns
            break
        prev_qtns = new_qtns
        pseudo_qtns = new_qtns

    return pvals, betas, ses

def _fixed_effect_scan(y, G, cov_idx, test_leave_one_out=True):
    """One fixed-effect-model scan of every marker in G, with the markers
    listed in cov_idx included as covariates (removed from the covariate
    set for a marker that IS one of them, i.e. tested leave-one-out).
    Shared by BLINK / mrMLM / FASTmrMLM below so their iteration logic
    stays readable."""
    n, m = G.shape
    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
    cov_set = set(cov_idx)
    for j in range(m):
        x = G[:, j]
        if x.std() < 1e-8:
            continue
        if test_leave_one_out and j in cov_set:
            others = [q for q in cov_idx if q != j]
        else:
            others = list(cov_idx)
        if others:
            cov = G[:, others] - G[:, others].mean(axis=0)
            Xj = np.column_stack([np.ones((n, 1)), cov])
        else:
            Xj = np.ones((n, 1))
        Xf = np.column_stack([Xj, x])
        try:
            beta = np.linalg.lstsq(Xf, y, rcond=None)[0]
            XtX_inv = np.linalg.pinv(Xf.T @ Xf)
        except np.linalg.LinAlgError:
            continue
        resid = y - Xf @ beta
        dfree = max(n - Xf.shape[1], 1)
        sigma2 = float(np.sum(resid ** 2) / dfree)
        se = float(np.sqrt(max(sigma2 * XtX_inv[-1, -1], 0)))
        b = float(beta[-1])
        t_stat = b / (se + 1e-300)
        pvals[j] = float(2 * stats.t.sf(abs(t_stat), df=dfree))
        betas[j] = b; ses[j] = se
    return pvals, betas, ses

def _bic_fixed_effect_model(y, G, cand_idx):
    """Fit y ~ intercept + G[:, cand_idx] and return BIC, used by BLINK to
    pick the best-supported pseudo-QTN subset instead of a fixed top-N."""
    n = len(y)
    if len(cand_idx) == 0:
        X = np.ones((n, 1))
    else:
        X = np.column_stack([np.ones((n, 1)), G[:, cand_idx]])
    beta, *_ = np.linalg.lstsq(X, y, rcond=None)
    resid = y - X @ beta
    k = X.shape[1]
    rss = float(np.sum(resid ** 2)) + 1e-300
    return n * np.log(rss / n) + k * np.log(n)

def blink_gwas(y, G, K, chroms=None, positions=None, max_iter=8, ld_r2=0.7,
                ld_block_bp=1_000_000):
    """
    BLINK (Huang et al. 2019, Bayesian-information and Linkage-disequilibrium
    Iteratively Nested Keyway): like FarmCPU, alternates a fixed-effect-model
    scan with pseudo-QTN selection, but differs in two places that give
    BLINK its name/speed advantage over FarmCPU: (1) candidate pseudo-QTNs
    are pruned within local LD blocks (candidates on the same chromosome and
    within ld_block_bp of an already-kept, more-significant candidate are
    compared by pairwise r^2 and dropped if redundant; candidates outside
    that window are never compared and are always kept, since LD is a local
    phenomenon and two markers hundreds of kb apart on different LD blocks
    can be "collapsed" by chance correlation in a small panel even with no
    real linkage) instead of a fixed physical bin size, so two markers 50bp
    apart in high LD are collapsed the same way two markers 500kb apart in
    low LD are kept separate; (2) the number of pseudo-QTNs carried into the
    next iteration is chosen by minimising BIC over nested candidate
    subsets, not a fixed top-10 cutoff. No kinship matrix is used once QTNs
    are found — BLINK, like FarmCPU, assumes the pseudo-QTNs themselves
    absorb population structure/relatedness.
    """
    n, m = G.shape
    if positions is None:
        positions = np.arange(m)
    if chroms is None:
        chroms = np.zeros(m, dtype=int)

    pseudo_qtns, prev_qtns = [], None
    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)

    for _ in range(max_iter):
        pvals, betas, ses = _fixed_effect_scan(y, G, pseudo_qtns)

        bonf = 0.05 / m
        cand = np.where(pvals < bonf)[0]
        if len(cand) == 0:
            cand = np.argsort(pvals)[:10]
        cand = cand[np.argsort(pvals[cand])]

        # LD-prune candidates: keep a candidate only if it isn't in high LD
        # with an already-kept, more-significant candidate FROM THE SAME LD
        # BLOCK (same chromosome, within ld_block_bp). Candidates outside
        # that window skip the correlation check entirely -- LD blocks are
        # local, so comparing markers genome-wide (the previous behaviour)
        # could collapse two unlinked-but-coincidentally-correlated distant
        # markers into one pseudo-QTN.
        kept = []
        for idx in cand:
            redundant = False
            for k_idx in kept:
                same_block = (chroms[idx] == chroms[k_idx] and
                              abs(int(positions[idx]) - int(positions[k_idx])) <= ld_block_bp)
                if not same_block:
                    continue
                r = np.corrcoef(G[:, idx], G[:, k_idx])[0, 1]
                if np.isfinite(r) and r ** 2 >= ld_r2:
                    redundant = True
                    break
            if not redundant:
                kept.append(idx)
            if len(kept) >= 15:
                break

        # BIC-based model selection: try nested prefixes of `kept`
        # (most-significant-first) and keep the prefix with lowest BIC
        best_bic, best_set = np.inf, []
        for size in range(0, len(kept) + 1):
            subset = kept[:size]
            bic = _bic_fixed_effect_model(y, G, subset)
            if bic < best_bic:
                best_bic, best_set = bic, subset
        new_qtns = best_set

        if prev_qtns is not None and set(new_qtns) == set(prev_qtns):
            pseudo_qtns = new_qtns
            break
        prev_qtns = new_qtns
        pseudo_qtns = new_qtns

    return pvals, betas, ses

def mrmlm_gwas(y, G, K, chroms=None, positions=None, screen_thresh=0.01,
               max_candidates=20, drop_thresh=0.05, n_pca=3):
    """
    mrMLM (multi-locus random-SNP-effect mixed linear model; Wang et al.
    2016): a two-step multi-locus approach. Step 1 runs a lenient
    single-locus P3D mixed-model scan (same engine as MLM/EMMAX) and keeps
    every marker below `screen_thresh` as a candidate QTN. Step 2 fits all
    candidates simultaneously as fixed effects (approximating mrMLM's
    random-SNP-effect / EM-Bayes shrinkage step) and backward-eliminates
    the least significant one at a time until every remaining candidate is
    significant at `drop_thresh` — mrMLM's real innovation is testing
    QTNs jointly rather than one-at-a-time, which is what's reproduced
    here. Markers never selected as candidates keep their Step-1 p-value
    so the full-length p-value vector still plots as a Manhattan track.
    """
    n, m = G.shape
    Q = get_kinship_pcs(K, n_pca, n)
    pvals, betas, ses, _ = mixed_model_gwas(y, G, K, covariates=Q)

    cand = np.where(pvals < screen_thresh)[0]
    if len(cand) == 0:
        return pvals, betas, ses
    cand = cand[np.argsort(pvals[cand])][:max_candidates].tolist()

    # backward elimination on the joint multi-locus fixed-effect model
    while len(cand) > 0:
        p_joint, b_joint, se_joint = _fixed_effect_scan(y, G, cand)
        cand_p = {c: p_joint[c] for c in cand}
        worst = max(cand_p, key=cand_p.get)
        if cand_p[worst] > drop_thresh and len(cand) > 1:
            cand.remove(worst)
            continue
        break

    p_final, b_final, se_final = _fixed_effect_scan(y, G, cand)
    for c in cand:
        pvals[c] = p_final[c]; betas[c] = b_final[c]; ses[c] = se_final[c]
    return pvals, betas, ses

def fastmrmlm_gwas(y, G, K, chroms=None, positions=None, screen_thresh=0.01,
                    max_candidates=20, drop_thresh=0.05, n_pca=3):
    """
    FASTmrMLM (Tamba et al. 2017): the speed-optimised successor to mrMLM.
    The real method reuses a single kinship eigen-decomposition (computed
    once) to transform y and G, avoiding a per-marker REML fit; that's
    exactly what EMMAX/mixed_model_gwas already does here, so Step 1
    candidate screening below IS that fast rotated scan. Step 2 departs
    from mrMLM by fitting the joint multi-locus model as plain OLS
    (dropping the kinship term for the joint step, trading a small amount
    of type-I-error control for the speed gain that gives the method its
    name) with the same backward-elimination logic as mrMLM.
    """
    n, m = G.shape
    Q = get_kinship_pcs(K, n_pca, n)
    pvals, betas, ses, _ = mixed_model_gwas(y, G, K, covariates=Q)

    cand = np.where(pvals < screen_thresh)[0]
    if len(cand) == 0:
        return pvals, betas, ses
    cand = cand[np.argsort(pvals[cand])][:max_candidates].tolist()

    while len(cand) > 0:
        p_joint, b_joint, se_joint = _fixed_effect_scan(y, G, cand,
                                                          test_leave_one_out=True)
        cand_p = {c: p_joint[c] for c in cand}
        worst = max(cand_p, key=cand_p.get)
        if cand_p[worst] > drop_thresh and len(cand) > 1:
            cand.remove(worst)
            continue
        break

    p_final, b_final, se_final = _fixed_effect_scan(y, G, cand,
                                                      test_leave_one_out=True)
    for c in cand:
        pvals[c] = p_final[c]; betas[c] = b_final[c]; ses[c] = se_final[c]
    return pvals, betas, ses

def gemma_gwas(y, G, K, covariates=None):
    """
    Approximate GEMMA-style univariate LMM (Zhou & Stephens 2012): unlike
    the P3D models above, the genetic/residual variance ratio is
    re-estimated PER MARKER (the candidate SNP is included in the null
    model before REML), which is GEMMA's default exact behaviour. Slower
    than P3D but the most statistically rigorous of the five — no post-hoc
    p-value rescaling is applied, since a correctly specified LMM shouldn't
    need genomic-control patching.
    """
    n, m = G.shape
    X0 = np.ones((n, 1))
    if covariates is not None and covariates.shape[1] > 0:
        X0 = np.column_stack([X0, covariates])
    evals, evecs = eigh_kinship(K)
    yt = evecs.T @ y
    X0t = evecs.T @ X0
    Gt = evecs.T @ G

    pvals = np.ones(m); betas = np.zeros(m); ses = np.zeros(m)
    for j in range(m):
        if G[:, j].std() < 1e-8:
            continue
        Xt = np.column_stack([X0t, Gt[:, j]])
        res = minimize_scalar(lambda ld: -_profile_loglik(ld, evals, yt, Xt, reml=True),
                               bounds=(-10, 10), method="bounded",
                               options={"xatol": 1e-3})
        delta = float(np.exp(res.x))
        p, b, se = _wald_test_rotated(yt, Xt, evals, delta)
        pvals[j] = p; betas[j] = b; ses[j] = se
    return pvals, betas, ses

# ─────────────────────── CONFIG BUILDER ──────────────────────────────────────────
def build_cfg(w, h, fs, fc, gc="#E8ECF4", bg="#FFFFFF", panel="#FAFBFF",
              sig_col="#E15759", sug_col="#4E79A7", marker_size=12, alpha=0.75,
              font_color=None, dot_col=None, line_col=None, fill_col=None,
              axis_col=None):
    """
    Build a plot configuration dictionary.
    'fc' is the required positional font-color argument.
    'font_color' is an optional alias (takes precedence if provided).
    """
    effective_fc = font_color if font_color is not None else fc
    return {
        "width":        w,
        "height":       h,
        "font_size":    fs,
        "font_color":   effective_fc,
        "grid_color":   gc,
        "bg_color":     bg,
        "panel_color":  panel,
        "sig_color":    sig_col,
        "sug_color":    sug_col,
        "marker_size":  marker_size,
        "alpha":        alpha,
        "dot_color":    dot_col  or "#4E79A7",
        "line_color":   line_col or "#E15759",
        "fill_color":   fill_col or "#A78BFA",
        "axis_color":   axis_col or "#2D3142",
    }

# ─────────────────────── MANHATTAN PLOT ──────────────────────────────────────────
def plot_manhattan(pvals, chroms, positions, title="Manhattan Plot", model_name="",
                   cfg=None, top_n=15, lod_threshold=None,
                   pve_vals=None, pve_threshold=None, show_logp_axis=True):
    """
    LOD is the selection criterion for this pipeline, so the LEFT axis
    always shows the LOD scale. The RIGHT axis is an optional secondary
    -log10(p) scale (show_logp_axis toggles it on/off) -- both axes are
    exact monotonic transforms of the same underlying data
    (neglog10p_to_lod / lod_to_neglog10p), not independent scales.
    """
    if cfg is None:
        cfg = build_cfg(16, 5, 9, "#2D3142")

    n_snps = len(pvals)
    bonf_thresh = 0.05 / n_snps
    sug_thresh  = 1.0  / n_snps

    sig_line = -np.log10(bonf_thresh)
    sug_line = -np.log10(sug_thresh)

    fig, ax = plt.subplots(figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])
    ax.set_facecolor(cfg["panel_color"])

    log_p = -np.log10(np.clip(pvals, 1e-300, 1))

    chrom_list = sorted(set(chroms))
    chrom_colors = {c: CHR_PALETTE[i % len(CHR_PALETTE)] for i, c in enumerate(chrom_list)}
    x_offset = 0
    tick_pos, tick_lab, x_coords = [], [], np.zeros(len(pvals))
    gap = max(positions) * 0.018 + 1

    for ch in chrom_list:
        mask = chroms == ch
        pos = positions[mask]
        if len(pos) == 0: continue
        order = np.argsort(pos)
        x_vals = pos[order] - pos.min() + x_offset
        x_coords[np.where(mask)[0][order]] = x_vals
        tick_pos.append(x_vals.mean())
        tick_lab.append(f"Chr{ch}")
        ax.scatter(x_coords[mask], log_p[mask],
                   c=chrom_colors[ch], s=cfg["marker_size"],
                   alpha=cfg["alpha"], linewidths=0)
        x_offset += (pos.max() - pos.min()) + gap

    ax.axhline(sug_line, color=cfg["sug_color"], ls=":", lw=1.5, alpha=0.9,
               label=f"Suggestive (1/m = {sug_thresh:.1e}, -log₁₀={sug_line:.1f})")
    ax.axhline(sig_line, color=cfg["sig_color"], ls="--", lw=2.0,
               label=f"Bonferroni (0.05/m = {bonf_thresh:.1e}, -log₁₀={sig_line:.1f})")
    if lod_threshold is not None:
        lod_line_negp = lod_to_neglog10p(lod_threshold)
        ax.axhline(lod_line_negp, color=ACCENT_LIME, ls="-", lw=2.0, alpha=0.95,
                   label=f"Permutation / LOD threshold (LOD={lod_threshold:.2f})")

    # PRIMARY significance criterion: markers passing the LOD/permutation
    # threshold (optionally AND a minimum PVE%), not a blind top-N slice.
    # If no threshold could be computed, falls back to a small top-N so the
    # plot doesn't render empty.
    sig_idx = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n,
                                   pve_vals=pve_vals, pve_threshold=pve_threshold)
    sig_mask = np.zeros(len(pvals), dtype=bool)
    sig_mask[sig_idx] = True

    if sig_mask.any():
        crit_bits = []
        if lod_threshold is not None:
            crit_bits.append("LOD threshold")
        if pve_vals is not None and pve_threshold:
            crit_bits.append(f"PVE ≥ {pve_threshold:.1f}%")
        label = (f"{sig_mask.sum()} MTA(s) ≥ {' & '.join(crit_bits)}" if crit_bits
                 else f"Top {sig_mask.sum()} MTAs (no threshold available)")
        ax.scatter(x_coords[sig_mask], log_p[sig_mask],
                   c=ACCENT_GOLD, s=max(cfg["marker_size"]*4, 60),
                   marker="D", zorder=10, edgecolors="#333333", linewidth=0.7,
                   label=label)

    ax.set_xticks(tick_pos)
    ax.set_xticklabels(tick_lab, rotation=45,
                       fontsize=cfg["font_size"]-1, color=cfg["font_color"])
    ax.set_xlabel("Chromosome", color=cfg["font_color"], fontsize=cfg["font_size"])
    ax.set_title(f"{title}{' — ' + model_name if model_name else ''}",
                 color=cfg["font_color"], fontsize=cfg["font_size"]+2, pad=8)
    ax.legend(fontsize=cfg["font_size"]-1, framealpha=0.9,
              facecolor=cfg["bg_color"], edgecolor="#CCCCCC",
              labelcolor=cfg["font_color"], loc='upper right')
    ax.set_xlim(-x_offset * 0.01, x_offset * 1.01)
    for spine in ax.spines.values():
        spine.set_edgecolor("#CCCCCC")
    ax.grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    # ── LEFT axis: the data is plotted on the -log10(p) scale, but every
    # tick is *labeled* in LOD units via a direct formatter -- this always
    # renders (unlike a stacked secondary axis, which can silently collapse
    # to zero width under tight_layout). LOD and -log10(p) are exact
    # monotonic transforms of each other, so this is just a relabeling.
    ax.yaxis.set_major_formatter(
        plt.FuncFormatter(lambda val, pos: f"{neglog10p_to_lod(val):.1f}")
    )
    ax.set_ylabel("LOD score", color=cfg["font_color"], fontsize=cfg["font_size"])
    ax.tick_params(axis='y', colors=cfg["font_color"], labelsize=cfg["font_size"]-1)

    if show_logp_axis:
        ax_logp = ax.secondary_yaxis('right')
        ax_logp.set_ylabel("-log₁₀(p)", color=cfg["font_color"], fontsize=cfg["font_size"])
        ax_logp.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        ax_logp.spines["right"].set_edgecolor("#CCCCCC")

    fig.tight_layout()
    return fig


# ─────────────────────── QQ PLOT ─────────────────────────────────────────────────
def _nature_qq_ticks(max_val):
    """Pick clean, evenly-spaced tick values (Nature-style QQ plots use round
    steps like 0,3,6,9,12 rather than whatever matplotlib defaults to)."""
    if not np.isfinite(max_val) or max_val <= 0:
        return [0, 1]
    for step in (1, 2, 3, 5, 10, 15, 20, 25, 50, 100):
        n_ticks = max_val / step
        if n_ticks <= 5:
            top = int(np.ceil(max_val / step)) * step
            return list(np.arange(0, top + step, step))
    top = int(np.ceil(max_val / 100)) * 100
    return list(np.arange(0, top + 100, 100))


def plot_qq_single(pvals, title="QQ Plot", model_name="", color=None, cfg=None,
                    panel_letter=None):
    """Clean single-series QQ plot in a minimal Nature-journal style: plain
    white background, only left/bottom axis lines, a thin black y=x
    reference line, no gridlines, and a small unobtrusive lambda label."""
    if cfg is None:
        cfg = build_cfg(7, 7, 9, "#000000")
    if color is None:
        color = cfg.get("dot_color", "#3E7BD6")

    fig, ax = plt.subplots(figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")

    n = len(pvals)
    p_sorted = np.sort(np.clip(pvals, 1e-300, 1))
    obs = -np.log10(p_sorted)
    exp = -np.log10((np.arange(1, n + 1) - 0.5) / n)

    ax.scatter(exp, obs, c=color, s=max(cfg["marker_size"] * 0.6, 8),
               alpha=0.9, linewidths=0, zorder=5)

    axis_max = float(np.nanmax([exp.max() if n else 1, obs.max() if n else 1]))
    ticks = _nature_qq_ticks(axis_max)
    lim = ticks[-1]
    ax.plot([0, lim], [0, lim], color="#000000", lw=1.3, zorder=4, solid_capstyle="round")

    lam = np.median(stats.chi2.ppf(1 - np.clip(pvals, 1e-10, 1), df=1)) / stats.chi2.ppf(0.5, df=1)
    ax.text(0.97, 0.04, f"\u03bb = {lam:.3f}", transform=ax.transAxes,
            fontsize=cfg["font_size"] - 2, color="#888888", ha="right", va="bottom")

    ax.set_xlim(0, lim)
    ax.set_ylim(0, lim)
    ax.set_xticks(ticks)
    ax.set_yticks(ticks)
    ax.set_xlabel("Expected \u2212log$_{10}$($P$ value)", fontsize=cfg["font_size"] + 1,
                  color="#000000")
    ax.set_ylabel("Observed \u2212log$_{10}$($P$ value)", fontsize=cfg["font_size"] + 1,
                  color="#000000")
    if title or model_name:
        ax.set_title(f"{title}{' — ' + model_name if model_name else ''}",
                     color="#555555", fontsize=cfg["font_size"], pad=10)
    if panel_letter:
        ax.text(-0.16, 1.04, panel_letter, transform=ax.transAxes,
                fontsize=cfg["font_size"] + 8, fontweight="bold", color="#000000",
                ha="left", va="top")

    ax.tick_params(colors="#000000", labelsize=cfg["font_size"], length=4, width=1.1)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color("#000000")
        ax.spines[side].set_linewidth(1.2)
    ax.grid(False)
    fig.tight_layout()
    return fig


def plot_qq_all_models(all_results, cfg=None):
    """QQ plots for all models with regression lines."""
    if cfg is None:
        cfg = build_cfg(7, 6, 9, "#2D3142")
    n_models = len(all_results)
    ncols = min(3, n_models)
    nrows = (n_models + ncols - 1) // ncols
    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(cfg["width"] * ncols, cfg["height"] * nrows))
    fig.patch.set_facecolor(cfg["bg_color"])

    if n_models == 1:
        axes = np.array([axes])
    axes = np.array(axes).flatten()

    for idx, (mname, (pvals, _, _)) in enumerate(all_results.items()):
        ax = axes[idx]
        ax.set_facecolor(cfg["panel_color"])
        n = len(pvals)
        p_sorted = np.sort(np.clip(pvals, 1e-300, 1))
        obs = -np.log10(p_sorted)
        exp = -np.log10((np.arange(1, n + 1) - 0.5) / n)
        ci_lo = -np.log10(stats.beta.ppf(0.975, np.arange(1, n+1), np.arange(n, 0, -1)))
        ci_hi = -np.log10(stats.beta.ppf(0.025, np.arange(1, n+1), np.arange(n, 0, -1)))
        col = MODEL_COLORS.get(mname, "#7C6FCD")

        ax.fill_between(exp, ci_lo, ci_hi, alpha=0.2, color=col)
        ax.scatter(exp, obs, c=col, s=cfg["marker_size"], alpha=cfg["alpha"],
                   linewidths=0, zorder=5)

        ax.plot([0, max(exp)], [0, max(exp)], color="#AAAAAA",
                ls="-", lw=1.0, alpha=0.6, zorder=4)

        valid = np.isfinite(exp) & np.isfinite(obs)
        if valid.sum() > 2:
            slope_reg, intercept_reg, r_val, _, _ = stats.linregress(exp[valid], obs[valid])
            x_reg = np.linspace(exp[valid].min(), exp[valid].max(), 200)
            y_reg = slope_reg * x_reg + intercept_reg
            ax.plot(x_reg, y_reg, color=cfg["sig_color"], ls="-", lw=2.0, zorder=6,
                    label=f"Reg slope={slope_reg:.2f}")
            ax.legend(fontsize=cfg["font_size"]-2, facecolor=cfg["bg_color"],
                      edgecolor="#CCCCCC", labelcolor=cfg["font_color"])

        lam = np.median(stats.chi2.ppf(1-np.clip(pvals,1e-10,1),df=1)) / stats.chi2.ppf(0.5, df=1)
        ax.text(0.05, 0.92, f"λ = {lam:.4f}", transform=ax.transAxes,
                fontsize=cfg["font_size"]+1, color="#E15759", fontweight='bold')
        ax.set_xlabel("Expected -log₁₀(p)", fontsize=cfg["font_size"], color=cfg["font_color"])
        ax.set_ylabel("Observed -log₁₀(p)", fontsize=cfg["font_size"], color=cfg["font_color"])
        ax.set_title(f"QQ — {mname}", color=col, fontsize=cfg["font_size"]+1, fontweight='bold')
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        for spine in ax.spines.values():
            spine.set_edgecolor("#CCCCCC")
        ax.grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    for idx in range(n_models, len(axes)):
        axes[idx].set_visible(False)

    fig.suptitle("QQ Plots — All Models (with Regression Lines)", color=cfg["font_color"],
                 fontsize=cfg["font_size"]+3, y=1.01, fontweight='bold')
    fig.tight_layout()
    return fig


def plot_qq_overlay(all_results, cfg=None, panel_letter=None):
    """Mixed multi-model QQ overlay styled after the Nature-journal
    convention: each model is a dense small-dot curve in its own color,
    a single thin black y=x line, and colored text labels placed right at
    the end of each curve instead of a boxed legend."""
    if cfg is None:
        cfg = build_cfg(8, 8, 11, "#000000")
    fig, ax = plt.subplots(figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")

    curves = {}
    axis_max = 0.0
    for mname, (pvals, _, _) in all_results.items():
        n = len(pvals)
        if n == 0:
            continue
        p_sorted = np.sort(np.clip(pvals, 1e-300, 1))
        obs = -np.log10(p_sorted)
        exp = -np.log10((np.arange(1, n + 1) - 0.5) / n)
        curves[mname] = (exp, obs)
        axis_max = max(axis_max, float(np.nanmax(exp)), float(np.nanmax(obs)))

    ticks = _nature_qq_ticks(axis_max)
    lim = ticks[-1]
    ax.plot([0, lim], [0, lim], color="#000000", lw=1.3, zorder=3, solid_capstyle="round")

    label_positions = []  # track placed label y-positions (axes coords) to avoid overlap
    for mname, (exp, obs) in curves.items():
        col = MODEL_COLORS.get(mname, "#3E7BD6")
        ax.scatter(exp, obs, s=max(cfg["marker_size"] * 0.45, 6), color=col,
                   alpha=0.9, linewidths=0, zorder=5)

        # Place the model name at the end of its own curve (top-right-ish),
        # nudged vertically so labels for different models don't collide.
        end_x, end_y = exp[-1], obs[-1]
        y_frac = np.clip(end_y / lim, 0.05, 0.97)
        while any(abs(y_frac - p) < 0.055 for p in label_positions):
            y_frac += 0.06
        label_positions.append(y_frac)
        ax.text(0.99, y_frac, mname, transform=ax.transAxes, color=col,
                fontsize=cfg["font_size"] + 1, fontweight="bold",
                ha="right", va="center",
                path_effects=[pe.withStroke(linewidth=3, foreground="#FFFFFF")])

    ax.set_xlim(0, lim)
    ax.set_ylim(0, lim)
    ax.set_xticks(ticks)
    ax.set_yticks(ticks)
    ax.set_xlabel("Expected \u2212log$_{10}$($P$ value)", fontsize=cfg["font_size"] + 1,
                  color="#000000")
    ax.set_ylabel("Observed \u2212log$_{10}$($P$ value)", fontsize=cfg["font_size"] + 1,
                  color="#000000")
    if panel_letter:
        ax.text(-0.16, 1.04, panel_letter, transform=ax.transAxes,
                fontsize=cfg["font_size"] + 8, fontweight="bold", color="#000000",
                ha="left", va="top")

    ax.tick_params(colors="#000000", labelsize=cfg["font_size"], length=4, width=1.1)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color("#000000")
        ax.spines[side].set_linewidth(1.2)
    ax.grid(False)
    fig.tight_layout()
    return fig


# ─────────────────────── LOD / -log10(p) CONVERSION HELPERS ──────────────────────
def neglog10p_to_lod(neglog10p):
    """Convert a -log10(p) value (1 d.f. test) to its equivalent LOD score.

    LOD = chi2 / (2*ln(10)), where chi2 is the 1-d.f. chi-square value
    whose upper-tail p-value equals 10**(-neglog10p). This is the
    standard linkage/QTL LOD scale, so a p-based threshold can be read
    directly as a LOD threshold as well.
    """
    neglog10p = np.atleast_1d(np.asarray(neglog10p, dtype=float))
    p = np.clip(10.0 ** (-neglog10p), 1e-300, 1.0)
    chi2_val = stats.chi2.isf(p, df=1)
    lod = chi2_val / (2.0 * np.log(10.0))
    return lod if lod.shape != (1,) else lod[0]

def lod_to_neglog10p(lod):
    """Exact inverse of neglog10p_to_lod: given a LOD score, return the
    -log10(p) value it corresponds to, so a LOD threshold can be drawn on
    a -log10(p)-scaled Manhattan/Circos axis."""
    lod = np.atleast_1d(np.asarray(lod, dtype=float))
    chi2_val = lod * 2.0 * np.log(10.0)
    p = stats.chi2.sf(chi2_val, df=1)
    neglog10p = -np.log10(np.clip(p, 1e-300, 1.0))
    return neglog10p if neglog10p.shape != (1,) else neglog10p[0]


# ─────────────────────── LD DECAY ────────────────────────────────────────────────
def plot_ld_decay(G, positions, cfg=None):
    if cfg is None:
        cfg = build_cfg(16, 6, 9, "#2D3142")
    fig, axes = plt.subplots(1, 2, figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])

    m = G.shape[1]
    dists_bp, r2s = [], []
    n_pairs = min(80, m)
    for i in range(n_pairs):
        for j in range(i+1, n_pairs):
            xi, xj = G[:, i], G[:, j]
            mask = ~(np.isnan(xi) | np.isnan(xj))
            if mask.sum() < 4: continue
            r, _ = stats.pearsonr(xi[mask], xj[mask])
            d = abs(positions[j] - positions[i])
            dists_bp.append(d); r2s.append(r**2)

    if not dists_bp:
        for ax in axes:
            ax.set_facecolor(cfg["panel_color"])
            ax.text(0.5, 0.5, "Insufficient data", ha='center',
                    transform=ax.transAxes, color=cfg["font_color"])
        fig.tight_layout()
        return fig

    dists_bp = np.array(dists_bp)
    dists_mb = dists_bp / 1_000_000
    r2s = np.array(r2s)

    def bin_ld(dists, r2s, n_bins=20):
        bins = np.percentile(dists, np.linspace(0, 100, n_bins + 1))
        centers, means, sems = [], [], []
        for k in range(len(bins)-1):
            mask = (dists >= bins[k]) & (dists < bins[k+1])
            if mask.sum() > 2:
                centers.append((bins[k] + bins[k+1]) / 2)
                means.append(np.mean(r2s[mask]))
                sems.append(stats.sem(r2s[mask]))
        return np.array(centers), np.array(means), np.array(sems)

    for i, (dist_arr, xlabel, title_str, dot_col, line_col) in enumerate([
        (dists_bp, "Distance (bp)", "LD Decay — Base Pairs (bp)",
         cfg.get("dot_color","#4E79A7"), cfg.get("line_color","#E15759")),
        (dists_mb, "Distance (Mb)", "LD Decay — Megabases (Mb)", "#59A14F", "#F28E2B"),
    ]):
        ax = axes[i]
        ax.set_facecolor(cfg["panel_color"])
        ax.scatter(dist_arr, r2s, c=dot_col, s=cfg["marker_size"],
                   alpha=cfg["alpha"], linewidths=0)
        bc, bm, bs = bin_ld(dist_arr, r2s)
        if len(bc) > 0:
            ax.plot(bc, bm, color=line_col, lw=2.5, marker='o',
                    ms=5, zorder=5, label="Binned mean r²")
            ax.fill_between(bc, bm - bs, bm + bs, alpha=0.18, color=line_col)
        ax.axhline(0.5, color="#EDC948", ls='--', lw=1.5, alpha=0.85, label="r² = 0.5")
        ax.axhline(0.2, color="#B07AA1", ls=':', lw=1.2, alpha=0.75, label="r² = 0.2")
        ax.set_xlabel(xlabel, fontsize=cfg["font_size"], color=cfg["font_color"])
        ax.set_ylabel("r² (LD)", fontsize=cfg["font_size"], color=cfg["font_color"])
        ax.set_title(title_str, color=cfg["font_color"], fontsize=cfg["font_size"]+1,
                     fontweight='bold')
        ax.legend(fontsize=cfg["font_size"]-1, framealpha=0.85,
                  facecolor=cfg["bg_color"], edgecolor="#CCCCCC",
                  labelcolor=cfg["font_color"])
        ax.set_ylim(0, 1.05)
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        for spine in ax.spines.values(): spine.set_edgecolor("#CCCCCC")
        ax.grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    fig.suptitle("Linkage Disequilibrium (LD) Decay Analysis",
                 color=cfg["font_color"], fontsize=cfg["font_size"]+3,
                 fontweight='bold')
    fig.tight_layout()
    return fig


# ─────────────────────── PCA (LIGHT THEME) ───────────────────────────────────────
def plot_pca(G, y, ids, cfg=None):
    if cfg is None:
        cfg = build_cfg(14, 10, 9, TEXT_DARK,
                        gc=GRID_LIGHT, bg=BG_LIGHT, panel=PANEL_LIGHT,
                        sig_col=ACCENT_TEAL_D, sug_col=ACCENT_GOLD_D,
                        marker_size=55, alpha=0.80)

    bg_col    = cfg.get("bg_color",    BG_LIGHT)
    panel_col = cfg.get("panel_color", PANEL_LIGHT)
    fc        = cfg.get("font_color",  TEXT_DARK)
    gc        = cfg.get("grid_color",  GRID_LIGHT)
    fs        = cfg.get("font_size",   9)

    fig, axes = plt.subplots(1, 3, figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(bg_col)
    axes = axes.flatten()

    G_sc = StandardScaler().fit_transform(G)
    n_comp = min(5, G_sc.shape[0]-1, G_sc.shape[1])
    pca = PCA(n_components=n_comp)
    pcs = pca.fit_transform(G_sc)
    var_exp = pca.explained_variance_ratio_ * 100

    for ax in axes:
        ax.set_facecolor(panel_col)
        for spine in ax.spines.values():
            spine.set_edgecolor(GRID_LIGHT)
        ax.grid(color=gc, linewidth=0.5, linestyle="--", alpha=0.7)
        ax.tick_params(colors=fc, labelsize=fs-1)

    cmap = "viridis"
    sc = axes[0].scatter(pcs[:,0], pcs[:,1], c=y, cmap=cmap,
                         s=cfg["marker_size"], edgecolors='white', linewidths=0.4, alpha=cfg["alpha"])
    axes[0].set_xlabel(f"PC1 ({var_exp[0]:.1f}%)", fontsize=fs, color=fc)
    axes[0].set_ylabel(f"PC2 ({var_exp[1]:.1f}%)", fontsize=fs, color=fc)
    axes[0].set_title("PC1 vs PC2", color=ACCENT_TEAL_D, fontsize=fs+1, fontweight='bold')
    cbar = plt.colorbar(sc, ax=axes[0])
    cbar.set_label("Phenotype", color=fc, fontsize=fs-1)
    cbar.ax.tick_params(colors=fc)
    cbar.ax.yaxis.set_tick_params(color=fc)
    plt.setp(cbar.ax.yaxis.get_ticklabels(), color=fc)

    if n_comp >= 3:
        sc2 = axes[1].scatter(pcs[:,1], pcs[:,2], c=y, cmap=cmap,
                              s=cfg["marker_size"], edgecolors='white', linewidths=0.4, alpha=cfg["alpha"])
        axes[1].set_xlabel(f"PC2 ({var_exp[1]:.1f}%)", fontsize=fs, color=fc)
        axes[1].set_ylabel(f"PC3 ({var_exp[2]:.1f}%)", fontsize=fs, color=fc)
        axes[1].set_title("PC2 vs PC3", color=ACCENT_GOLD_D, fontsize=fs+1, fontweight='bold')
        cbar2 = plt.colorbar(sc2, ax=axes[1])
        cbar2.set_label("Phenotype", color=fc, fontsize=fs-1)
        cbar2.ax.tick_params(colors=fc)
        plt.setp(cbar2.ax.yaxis.get_ticklabels(), color=fc)
    else:
        axes[1].set_visible(False)

    sc3 = axes[2].scatter(pcs[:,0], pcs[:,1], c=y, cmap="plasma",
                          s=cfg["marker_size"], edgecolors='white', linewidths=0.4, alpha=cfg["alpha"])
    axes[2].set_title("Population Structure", color=ACCENT_PINK_D, fontsize=fs+1, fontweight='bold')
    axes[2].set_xlabel(f"PC1 ({var_exp[0]:.1f}%)", fontsize=fs, color=fc)
    axes[2].set_ylabel(f"PC2 ({var_exp[1]:.1f}%)", fontsize=fs, color=fc)
    cbar3 = plt.colorbar(sc3, ax=axes[2])
    cbar3.set_label("Phenotype", color=fc, fontsize=fs-1)
    cbar3.ax.tick_params(colors=fc)
    plt.setp(cbar3.ax.yaxis.get_ticklabels(), color=fc)

    fig.suptitle("Principal Component Analysis", color=TEXT_DARK,
                 fontsize=fs+3, fontweight='bold')
    fig.tight_layout()
    return fig


# ─────────────────────── MODEL COMPARISON MANHATTAN ──────────────────────────────
def plot_model_comparison_manhattan(all_results, chroms, positions, cfg=None, top_n=15, lod_threshold=None):
    if cfg is None:
        cfg = build_cfg(16, 3.5, 7, "#2D3142")
    n_models = len(all_results)
    fig, axes = plt.subplots(n_models, 1,
                             figsize=(cfg["width"], cfg["height"] * n_models),
                             sharex=True)
    fig.patch.set_facecolor(cfg["bg_color"])
    if n_models == 1:
        axes = [axes]

    chrom_list = sorted(set(chroms))
    total_len = 0
    chrom_offsets = {}
    for ch in chrom_list:
        chrom_offsets[ch] = total_len
        mask = chroms == ch
        if mask.sum() > 0:
            total_len += positions[mask].max() - positions[mask].min() + max(positions) * 0.02 + 1

    x_coords = np.zeros(len(chroms))
    for ch in chrom_list:
        mask = chroms == ch
        if mask.sum() > 0:
            x_coords[mask] = positions[mask] - positions[mask].min() + chrom_offsets[ch]

    tick_pos, tick_lab = [], []
    for ch in chrom_list:
        mask = chroms == ch
        if mask.sum() > 0:
            tick_pos.append(x_coords[mask].mean())
            tick_lab.append(f"Chr{ch}")

    for midx, (mname, (pvals, _, _)) in enumerate(all_results.items()):
        ax = axes[midx]
        ax.set_facecolor(cfg["panel_color"])
        log_p = -np.log10(np.clip(pvals, 1e-300, 1))
        col = MODEL_COLORS.get(mname, CHR_PALETTE[midx % len(CHR_PALETTE)])

        n_snps = len(pvals)
        bonf_thresh = 0.05 / n_snps
        sug_thresh  = 1.0  / n_snps

        for ch in chrom_list:
            mask = chroms == ch
            ax.scatter(x_coords[mask], log_p[mask],
                       c=col, s=cfg["marker_size"], alpha=cfg["alpha"], linewidths=0)

        sig_idx = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n)
        label = (f"MTA(s) ≥ LOD threshold ({len(sig_idx)})" if lod_threshold is not None
                 else f"Top {len(sig_idx)} (no threshold available)")
        if len(sig_idx) > 0:
            ax.scatter(x_coords[sig_idx], log_p[sig_idx],
                       c=ACCENT_GOLD, s=max(cfg["marker_size"]*4, 35),
                       zorder=10, linewidths=0.7, edgecolors='#333333',
                       marker='D', label=label)

        if lod_threshold is not None:
            ax.axhline(lod_to_neglog10p(lod_threshold), color=ACCENT_LIME,
                       ls='-', lw=1.6, alpha=0.95,
                       label=f"Permutation/LOD ({lod_threshold:.2f})")
        ax.axhline(-np.log10(sug_thresh), color=cfg["sug_color"],
                   ls=':', lw=1.2, alpha=0.9,
                   label=f"Sug. (1/m={sug_thresh:.1e})")
        ax.axhline(-np.log10(bonf_thresh), color=cfg["sig_color"],
                   ls='--', lw=1.5, alpha=0.9,
                   label=f"Bonf. (0.05/m={bonf_thresh:.1e})")

        ax.set_ylabel("-log₁₀(p)", fontsize=cfg["font_size"], color=cfg["font_color"])
        ax.set_title(f"Model: {mname}", color=col,
                     fontsize=cfg["font_size"]+1, pad=3, fontweight='bold')
        ax.set_xticks(tick_pos)
        ax.set_xticklabels(
            tick_lab if midx == n_models-1 else [],
            rotation=40, fontsize=cfg["font_size"]-1, color=cfg["font_color"]
        )
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        ax.legend(fontsize=cfg["font_size"]-2, framealpha=0.85,
                  facecolor=cfg["bg_color"], edgecolor="#CCCCCC",
                  labelcolor=cfg["font_color"], ncol=3)
        for spine in ax.spines.values(): spine.set_edgecolor("#CCCCCC")
        ax.grid(color=cfg["grid_color"], linewidth=0.3, linestyle="--", alpha=0.5)

    fig.suptitle("Manhattan Plots — All Models Comparison",
                 color=cfg["font_color"], fontsize=cfg["font_size"]+3,
                 fontweight='bold')
    fig.tight_layout(rect=[0, 0, 1, 0.98])
    return fig


# ─────────────────────── KINSHIP ─────────────────────────────────────────────────
def plot_kinship(K, ids, cfg=None):
    if cfg is None:
        cfg = build_cfg(14, 12, 9, "#2D3142")

    fig, axes = plt.subplots(2, 2, figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])
    axes = axes.flatten()
    for ax in axes:
        ax.set_facecolor(cfg["panel_color"])
        for spine in ax.spines.values(): spine.set_edgecolor("#CCCCCC")
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)

    im = axes[0].imshow(K, cmap='RdYlBu_r', aspect='auto', vmin=-0.5, vmax=1)
    axes[0].set_title("Kinship Matrix", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')
    cbar = plt.colorbar(im, ax=axes[0])
    cbar.ax.tick_params(colors=cfg["font_color"])

    K_flat = K[np.triu_indices_from(K, k=1)]
    axes[1].hist(K_flat, bins=30, color=cfg.get("dot_color","#4E79A7"),
                 edgecolor='white', alpha=0.85)
    axes[1].set_title("Pairwise Kinship Distribution", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[1].set_xlabel("Kinship Coefficient", fontsize=cfg["font_size"],
                       color=cfg["font_color"])
    axes[1].set_ylabel("Count", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[1].grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    try:
        lm = linkage(K, method='ward')
        n_samples = len(ids)
        leaf_fs = max(3, min(cfg["font_size"]-1, int(200 / max(n_samples, 1))))
        truncate_mode = None
        p_param = None
        if n_samples > 50:
            truncate_mode = 'lastp'
            p_param = 30
            axes[2].set_title(
                f"Dendrogram (truncated, {n_samples} samples)",
                color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold'
            )
        else:
            axes[2].set_title("Clustering Dendrogram", color=cfg["font_color"],
                              fontsize=cfg["font_size"]+1, fontweight='bold')

        dend_kwargs = dict(
            ax=axes[2],
            leaf_rotation=90,
            leaf_font_size=leaf_fs,
            color_threshold=0.7 * max(lm[:, 2]),
            above_threshold_color="#AAAAAA"
        )
        if truncate_mode:
            dend_kwargs["truncate_mode"] = truncate_mode
            dend_kwargs["p"] = p_param
        else:
            short_labels = [str(x)[:10] for x in ids]
            dend_kwargs["labels"] = short_labels

        dendrogram(lm, **dend_kwargs)
        axes[2].tick_params(axis='x', labelsize=leaf_fs, colors=cfg["font_color"])
        axes[2].tick_params(axis='y', labelsize=cfg["font_size"]-2,
                            colors=cfg["font_color"])
        if n_samples > 50:
            axes[2].set_xticks([])
        axes[2].set_facecolor(cfg["panel_color"])
        for spine in axes[2].spines.values(): spine.set_edgecolor("#CCCCCC")
    except Exception as e:
        axes[2].text(0.5, 0.5, f"Dendrogram\nunavailable:\n{str(e)[:40]}",
                     ha='center', va='center', transform=axes[2].transAxes,
                     color=cfg["font_color"], fontsize=cfg["font_size"])

    eigvals, eigvecs = np.linalg.eigh(K)
    idx = np.argsort(eigvals)[::-1]
    axes[3].scatter(eigvecs[:,idx[0]], eigvecs[:,idx[1]],
                    c=cfg.get("dot_color","#4E79A7"), s=cfg["marker_size"],
                    edgecolors='none', alpha=cfg["alpha"])
    axes[3].set_title("Kinship PCA", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[3].set_xlabel("PC1", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[3].set_ylabel("PC2", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[3].grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    fig.suptitle("Kinship Matrix & Clustering Analysis",
                 color=cfg["font_color"], fontsize=cfg["font_size"]+3, fontweight='bold')
    fig.tight_layout()
    return fig


# ─────────────────────── EFFECT SIZES ────────────────────────────────────────────
def plot_effect_sizes(pvals, betas, snp_names, cfg=None, top_n=15, lod_threshold=None,
                       pve_vals=None, pve_threshold=None):
    if cfg is None:
        cfg = build_cfg(14, 10, 9, "#2D3142")

    fig, axes = plt.subplots(2, 2, figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])
    axes = axes.flatten()
    for ax in axes:
        ax.set_facecolor(cfg["panel_color"])
        for spine in ax.spines.values(): spine.set_edgecolor("#CCCCCC")
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        ax.grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    top_idx = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n,
                                   pve_vals=pve_vals, pve_threshold=pve_threshold)
    n_shown = len(top_idx)
    top_betas = betas[top_idx]
    top_names = [str(snp_names[i])[:18] for i in top_idx]
    pos_col = cfg.get("dot_color", "#4E79A7")
    neg_col = cfg.get("line_color", "#E15759")
    colors = [neg_col if b < 0 else pos_col for b in top_betas]

    order = np.argsort(top_betas)
    axes[0].barh([top_names[o] for o in order], [top_betas[o] for o in order],
                 color=[colors[o] for o in order], edgecolor='white', linewidth=0.4)
    axes[0].axvline(0, color=cfg["font_color"], lw=0.9, alpha=0.6)
    axes[0].set_xlabel("Effect Size (β)", fontsize=cfg["font_size"],
                       color=cfg["font_color"])
    axes[0].set_title(
        f"Effect Sizes — {n_shown} MTA(s) ≥ LOD threshold" if lod_threshold is not None
        else f"Top {n_shown} Effect Sizes (no threshold available)",
        color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[0].tick_params(axis='y', labelsize=max(5, cfg["font_size"]-2))
    if n_shown == 0:
        axes[0].text(0.5, 0.5, "No MTAs passed the\npermutation/LOD threshold",
                    ha='center', va='center', transform=axes[0].transAxes,
                    color=cfg["font_color"], fontsize=cfg["font_size"])

    log_p = -np.log10(np.clip(pvals, 1e-300, 1))
    sig_thr = -np.log10(0.05/len(pvals))
    pt_colors = [neg_col if lp >= sig_thr else (pos_col if b > 0 else "#59A14F")
                 for lp, b in zip(log_p, betas)]
    axes[1].scatter(betas, log_p, c=pt_colors, s=cfg["marker_size"],
                    alpha=cfg["alpha"], linewidths=0)
    axes[1].axhline(sig_thr, color=cfg["sig_color"], ls='--', lw=1.8, label="Bonferroni")
    if lod_threshold is not None:
        axes[1].axhline(lod_to_neglog10p(lod_threshold), color=ACCENT_LIME,
                        ls='-', lw=1.8, label=f"Permutation/LOD ({lod_threshold:.2f})")
    axes[1].axvline(0, color=cfg["font_color"], lw=0.6, alpha=0.5)
    axes[1].set_xlabel("β", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[1].set_ylabel("-log₁₀(p)", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[1].set_title("Volcano Plot", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[1].legend(fontsize=cfg["font_size"]-1, facecolor=cfg["bg_color"],
                   edgecolor="#CCCCCC", labelcolor=cfg["font_color"])

    axes[2].hist(betas, bins=25, color=cfg.get("fill_color","#A78BFA"),
                 edgecolor='white', alpha=0.85)
    axes[2].axvline(0, color=cfg["sig_color"], lw=2, ls='--')
    axes[2].set_xlabel("β", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[2].set_title("Effect Size Distribution", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[2].set_ylabel("Count", fontsize=cfg["font_size"], color=cfg["font_color"])

    axes[3].scatter(betas, log_p, c=pt_colors, s=cfg["marker_size"],
                    alpha=cfg["alpha"]*0.85, linewidths=0)
    label_idx = top_idx[:min(8, len(top_idx))]
    for ti in label_idx:
        axes[3].annotate(str(snp_names[ti])[:12],
                         (betas[ti], log_p[ti]),
                         fontsize=max(5, cfg["font_size"]-3),
                         color=cfg["font_color"], alpha=0.8,
                         xytext=(5, 3), textcoords='offset points',
                         arrowprops=dict(arrowstyle='->', color='gray', lw=0.6))
    axes[3].set_xlabel("β", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[3].set_ylabel("-log₁₀(p)", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[3].set_title("Effect Size vs Significance (annotated)",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')

    fig.suptitle("Effect Size Analysis", color=cfg["font_color"],
                 fontsize=cfg["font_size"]+3, fontweight='bold')
    fig.tight_layout()
    return fig


# ─────────────────────── SUMMARY DASHBOARD ───────────────────────────────────────
def plot_summary_dashboard(y, maf, all_results, G_filt, cfg=None):
    if cfg is None:
        cfg = build_cfg(15, 10, 9, "#2D3142")

    fig, axes = plt.subplots(2, 3, figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])
    axes = axes.flatten()
    for ax in axes:
        ax.set_facecolor(cfg["panel_color"])
        for spine in ax.spines.values(): spine.set_edgecolor("#CCCCCC")
        ax.tick_params(colors=cfg["font_color"], labelsize=cfg["font_size"]-1)
        ax.grid(color=cfg["grid_color"], linewidth=0.4, linestyle="--", alpha=0.6)

    # Panel 0: Phenotype distribution
    axes[0].hist(y, bins=15, color=cfg.get("dot_color","#4E79A7"),
                 edgecolor='white', alpha=0.85)
    axes[0].axvline(y.mean(), color=cfg["sig_color"], lw=2, ls='--',
                    label=f"Mean: {y.mean():.2f}")
    axes[0].set_title(f"Phenotype Distribution | n={len(y)}",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[0].legend(fontsize=cfg["font_size"]-1, facecolor=cfg["bg_color"],
                   edgecolor="#CCCCCC", labelcolor=cfg["font_color"])
    axes[0].set_xlabel("Phenotype Value", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[0].set_ylabel("Count", fontsize=cfg["font_size"], color=cfg["font_color"])

    # Panel 1: MAF distribution
    axes[1].hist(maf, bins=20, color="#59A14F", edgecolor='white', alpha=0.85)
    axes[1].axvline(0.05, color=cfg["sig_color"], ls='--', lw=1.5,
                    label=f"MAF=0.05 (rare: {np.sum(maf<0.05)})")
    axes[1].set_title("Minor Allele Frequency",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[1].legend(fontsize=cfg["font_size"]-1, facecolor=cfg["bg_color"],
                   edgecolor="#CCCCCC", labelcolor=cfg["font_color"])
    axes[1].set_xlabel("MAF", fontsize=cfg["font_size"], color=cfg["font_color"])
    axes[1].set_ylabel("Count", fontsize=cfg["font_size"], color=cfg["font_color"])

    # Panel 2: Genomic inflation per model
    lambdas = {}
    for mname, (p, _, _) in all_results.items():
        lam = np.median(stats.chi2.ppf(1-np.clip(p,1e-10,1),df=1)) / stats.chi2.ppf(0.5,df=1)
        lambdas[mname] = lam
    lam_cols = ["#59A14F" if v<=1.1 else "#F28E2B" if v<=1.3 else "#E15759"
                for v in lambdas.values()]
    bars = axes[2].bar(lambdas.keys(), lambdas.values(),
                       color=lam_cols, edgecolor='white', linewidth=0.5)
    axes[2].axhline(1.0, color=cfg["font_color"], ls='--', lw=1, alpha=0.5)
    axes[2].set_title("Genomic Inflation λ per Model",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[2].tick_params(axis='x', rotation=30, labelsize=cfg["font_size"])
    axes[2].set_ylabel("λ", fontsize=cfg["font_size"], color=cfg["font_color"])
    for bar, lv in zip(bars, lambdas.values()):
        axes[2].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
                     f"{lv:.3f}", ha='center', fontsize=cfg["font_size"]-2,
                     color=cfg["font_color"], va='bottom')

    # Panel 3: Significant SNPs (Bonferroni)
    sig_counts = {}
    for mname, (p, _, _) in all_results.items():
        bonf = 0.05 / len(p)
        sig_counts[mname] = int((p < bonf).sum())
    cols_bar = [MODEL_COLORS.get(m, "#4E79A7") for m in sig_counts.keys()]
    axes[3].bar(sig_counts.keys(), sig_counts.values(),
                color=cols_bar, edgecolor='white', linewidth=0.5)
    axes[3].set_title("Significant SNPs (Bonferroni)",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')
    axes[3].tick_params(axis='x', rotation=30, labelsize=cfg["font_size"])
    axes[3].set_ylabel("Count", fontsize=cfg["font_size"], color=cfg["font_color"])

    # Panel 4: Model concordance heatmap
    n_m = len(all_results)
    model_names = list(all_results.keys())
    concordance = np.eye(n_m)
    for i in range(n_m):
        for j in range(i+1, n_m):
            lp_i = -np.log10(np.clip(all_results[model_names[i]][0], 1e-300, 1))
            lp_j = -np.log10(np.clip(all_results[model_names[j]][0], 1e-300, 1))
            r, _ = stats.spearmanr(lp_i, lp_j)
            concordance[i,j] = concordance[j,i] = r
    im = axes[4].imshow(concordance, cmap='RdYlGn', vmin=-1, vmax=1, aspect='auto')
    axes[4].set_xticks(range(n_m)); axes[4].set_yticks(range(n_m))
    axes[4].set_xticklabels(model_names, rotation=35,
                             fontsize=cfg["font_size"]-1, color=cfg["font_color"])
    axes[4].set_yticklabels(model_names, fontsize=cfg["font_size"]-1,
                             color=cfg["font_color"])
    for i in range(n_m):
        for j in range(n_m):
            axes[4].text(j, i, f"{concordance[i,j]:.2f}",
                         ha='center', va='center',
                         fontsize=cfg["font_size"]-1,
                         color='#111111' if abs(concordance[i,j])<0.7 else '#FFFFFF')
    cbar2 = plt.colorbar(im, ax=axes[4])
    cbar2.ax.tick_params(colors=cfg["font_color"])
    axes[4].set_title("Model Concordance (Spearman ρ)",
                      color=cfg["font_color"], fontsize=cfg["font_size"]+1, fontweight='bold')

    # Panel 5: Threshold table per model (replaces scree plot)
    axes[5].axis('off')
    thresh_lines = [
        f"{'Model':<10} {'Bonferroni':>14} {'Suggestive':>14}",
        "─" * 42,
    ]
    for mname, (p, _, _) in all_results.items():
        n_snps = len(p)
        b_thr = 0.05 / n_snps
        s_thr = 1.0  / n_snps
        thresh_lines.append(
            f"{mname:<10} {b_thr:>14.3e} {s_thr:>14.3e}"
        )
    thresh_lines.append("─" * 42)
    thresh_lines.append(f"  (Bonferroni = 0.05 / m)")
    thresh_lines.append(f"  (Suggestive =  1.0 / m)")

    summary_text = (
        f"GWAS Summary\n\n"
        f"Samples:       {G_filt.shape[0]}\n"
        f"SNPs (filt.):  {G_filt.shape[1]}\n"
        f"Models run:    {n_m}\n"
        f"Missing rate:  {np.isnan(G_filt).mean():.1%}\n\n"
        + "\n".join(thresh_lines)
    )
    axes[5].text(0.04, 0.96, summary_text,
                 ha='left', va='top', transform=axes[5].transAxes,
                 color=cfg["font_color"], fontsize=cfg["font_size"]-1,
                 fontfamily='monospace',
                 bbox=dict(boxstyle='round,pad=0.8', facecolor=cfg["panel_color"],
                           edgecolor="#CCCCCC", alpha=0.9))
    axes[5].set_title("Pipeline Summary + Thresholds", color=cfg["font_color"],
                      fontsize=cfg["font_size"]+1, fontweight='bold')

    fig.suptitle("GWAS Analysis Dashboard",
                 color=cfg["font_color"], fontsize=cfg["font_size"]+4, fontweight='bold')
    fig.tight_layout()
    return fig


# ─────────────────────── CIRCOS — DENSITY RING (pyCirclize) ──────────────────────
# Rebuilt on pyCirclize (github.com/moshi4/pyCirclize) instead of hand-rolled
# matplotlib polar geometry. The old approach manually computed every radian,
# bin edge, and fill polygon, which is exactly what produced the stray-line
# rendering artifact (a degenerate fill polygon collapsing into a spike).
# pyCirclize handles sector sizing, tracks, ticks, and heatmaps internally,
# so that whole class of bug is structurally no longer possible here.
CIRCOS_THEMES = {
    "Bright": {
        "bg": "#FFFFFF", "text": "#2B2B2B", "grid": "#DADADA",
        "chr_border": "#3A3A3A", "sig": "#C0392B", "star_edge": "#6B1A12",
        "track_colors": ["#2F6F9F", "#D98324", "#3E8E7E", "#8B5FBF", "#B5541A"],
        "density_stops": ["#F2F2F2", "#7FB77E", "#F4D35E", "#EE964B", "#C0392B"],
    },
    "Dark": {
        "bg": "#12141F", "text": "#F1F3FA", "grid": "#333A52",
        "chr_border": "#AEB6D6", "sig": "#FF6B6B", "star_edge": "#FFD9D9",
        "track_colors": ["#FF6B5B", "#FFD166", "#4FE0CB", "#D9A6F0", "#F2946B"],
        "density_stops": ["#2A2E45", "#3EC46D", "#F4E04D", "#FFA53E", "#FF4D4D"],
    },
}


def plot_circos_density(all_results, chroms, positions, ring_names, theme_name="Bright",
                        rings_per_panel=3, cfg=None, lod_threshold=None):
    """Build a single publication-quality circular GWAS figure (pyCirclize).

    Kept the original call signature (all_results, chroms, positions, ring_names,
    theme_name, rings_per_panel, cfg, lod_threshold) so run_gwas()'s call site and
    the "All Plots as PNG (ZIP)" export need no other changes.

    `rings_per_panel` is accepted for backward compatibility with the Gradio
    slider, but no longer splits the figure into an A/B/C/D subplot grid: all
    requested rings are drawn as concentric tracks in one circle, which is the
    standard circlize/Circos convention for a submission-ready figure. If you
    pass many models at once, consider trimming `ring_names` (the "Circos
    models" selector in the UI) rather than relying on paneling to declutter.
    """
    theme = CIRCOS_THEMES.get(theme_name, CIRCOS_THEMES["Bright"])
    fig_size = (cfg.get("width", 20) if cfg else 20) / 1.8
    font_size = cfg.get("font_size", 11) if cfg else 11
    density_bins_per_chr = 60
    manhattan_bins_per_chr = 200

    dens_cmap = LinearSegmentedColormap.from_list("density", theme["density_stops"], N=256)

    chroms = np.asarray(chroms)
    positions = np.asarray(positions, dtype=float)
    chr_list = sorted(set(int(c) for c in chroms))

    # Normalize each chromosome's positions to start at 0 (subtract its own
    # min), exactly like plot_manhattan() does. Without this, any position
    # source that doesn't start near 0 for every chromosome (e.g. the
    # per-chromosome fallback index, or real bp coordinates whose first
    # marker isn't near the telomere) makes chr_lengths[c] (= raw max)
    # larger than the marker's true span, leaving a genuinely empty arc at
    # the start of that sector.
    chr_min = {c: float(np.min(positions[chroms == c])) for c in chr_list}
    positions = positions.copy()
    for c in chr_list:
        mask = chroms == c
        positions[mask] = positions[mask] - chr_min[c]

    chr_lengths = {c: max(float(np.max(positions[chroms == c])), 1.0) for c in chr_list}
    sectors = {str(c): chr_lengths[c] for c in chr_list}

    circos = Circos(sectors, space=2.0)

    # ── Track 1: chromosome ideogram (outermost) ───────────────────────
    for sector in circos.sectors:
        c = int(sector.name)
        ideogram = sector.add_track((97, 100))
        ideogram.axis(fc=theme["bg"], ec=theme["chr_border"], lw=1.0)
        ideogram.text(f"{c}", size=font_size + 1, color=theme["text"], weight="bold")
        span_mb = chr_lengths[c] / 1e6
        interval = 50 if span_mb > 150 else (20 if span_mb > 50 else 10)
        ideogram.xticks_by_interval(
            interval * 1e6,
            label_formatter=lambda v: f"{v / 1e6:.0f}",
            label_size=font_size - 4, label_orientation="vertical", tick_length=1.0,
            line_kws=dict(color=theme["grid"], lw=0.6), text_kws=dict(color=theme["text"]),
        )

    # ── Track 2: SNP density heatmap ────────────────────────────────────
    dens_vals_all = []
    for sector in circos.sectors:
        c = int(sector.name)
        p = positions[chroms == c]
        bins = np.linspace(0, chr_lengths[c], density_bins_per_chr + 1)
        counts, _ = np.histogram(p, bins=bins)
        dens_vals_all.append(counts)
    dens_vmax = max((int(np.max(c)) if len(c) else 1 for c in dens_vals_all), default=1)
    dens_vmax = max(dens_vmax, 1)

    for sector, counts in zip(circos.sectors, dens_vals_all):
        dtrack = sector.add_track((90, 96))
        dtrack.axis(fc=theme["bg"], ec=theme["chr_border"], lw=0.6)
        dtrack.heatmap(counts.reshape(1, -1), vmin=0, vmax=dens_vmax, cmap=dens_cmap,
                       rect_kws=dict(ec="none"))

    # ── Tracks 3..N: one Manhattan ring per model ───────────────────────
    n_rings = max(len(ring_names), 1)
    ring_span, ring_gap = 82.0, 3.0
    each = (ring_span - ring_gap * (n_rings - 1)) / n_rings
    top_r = 90.0

    legend_handles = []
    for ridx, name in enumerate(ring_names):
        if name not in all_results:
            continue
        pvals = np.asarray(all_results[name][0])
        color = theme["track_colors"][ridx % len(theme["track_colors"])]
        log_p = -np.log10(np.clip(pvals, 1e-300, 1))
        lod_disp = lod_threshold if lod_threshold is not None else float(np.nanpercentile(log_p, 99.5))
        y_max = max(np.ceil(max(float(np.nanmax(log_p)), lod_disp + 1)), lod_disp + 1)

        r_top = top_r - ridx * (each + ring_gap)
        r_bot = r_top - each

        for si, sector in enumerate(circos.sectors):
            c = int(sector.name)
            idx = np.where(chroms == c)[0]
            if idx.size == 0:
                continue
            p_ch, lp_ch = positions[idx], log_p[idx]

            track = sector.add_track((r_bot, r_top))
            track.axis(fc=theme["bg"], ec=theme["grid"], lw=0.4)
            track.grid(y_grid_num=3, color=theme["grid"], lw=0.4, ls=(0, (1, 2)))

            if si == 0:
                track.yticks(
                    [0, round(lod_disp, 1), round(y_max)], vmin=0, vmax=y_max,
                    side="left", tick_length=1.2, label_size=font_size - 4.5,
                    label_margin=1.0, text_kws=dict(color=theme["text"]),
                    line_kws=dict(color=theme["grid"], lw=0.5),
                )

            n_bins = max(int(manhattan_bins_per_chr * (chr_lengths[c] / max(chr_lengths.values()))), 8)
            edges = np.linspace(0, chr_lengths[c], n_bins + 1)
            centers = (edges[:-1] + edges[1:]) / 2
            bin_idx = np.clip(np.digitize(p_ch, edges) - 1, 0, n_bins - 1)
            bin_max = np.zeros(n_bins)
            for bi in range(n_bins):
                sel = bin_idx == bi
                if sel.any():
                    bin_max[bi] = lp_ch[sel].max()

            has_signal = bin_max > 0
            if has_signal.any():
                track.bar(centers[has_signal], bin_max[has_signal],
                          width=(edges[1] - edges[0]) * 0.95, vmin=0, vmax=y_max,
                          color=color, ec="none")

            track.line([0, chr_lengths[c]], [lod_disp, lod_disp], vmin=0, vmax=y_max,
                      color=theme["sig"], lw=1.0, ls=(0, (1.5, 1.5)))

            passes = lp_ch >= lod_disp
            if passes.any():
                track.scatter(p_ch[passes], lp_ch[passes], vmin=0, vmax=y_max,
                             marker="*", s=28, color=theme["sig"],
                             ec=theme["star_edge"], lw=0.4, zorder=5)

        legend_handles.append(Patch(facecolor=color, edgecolor="none", label=name))

    legend_handles.append(Line2D([0], [0], color=theme["sig"], lw=1.3, ls=(0, (1.5, 1.5)),
                                  label="Significance threshold"))

    fig = circos.plotfig(figsize=(fig_size, fig_size))
    fig.patch.set_facecolor(theme["bg"])

    circos.colorbar(bounds=(1.00, 0.72, 0.02, 0.22), vmin=0, vmax=dens_vmax, cmap=dens_cmap,
                     label="SNP density\n(count / bin)",
                     label_kws=dict(size=font_size - 2, color=theme["text"]),
                     tick_kws=dict(labelsize=font_size - 3, colors=theme["text"]))

    fig.legend(handles=legend_handles, loc="center left", bbox_to_anchor=(0.98, 0.35),
               frameon=False, fontsize=font_size - 1, labelcolor=theme["text"],
               handlelength=1.4, borderaxespad=0.0, title="Tracks", title_fontsize=font_size)

    return fig




# ─────────────────────── EXCEL EXPORT ────────────────────────────────────────────
def export_results_excel(all_results, snp_filt, chr_filt, pos_filt, y,
                         G_filt, maf_vals, output_path, top_n=15, lod_threshold=None,
                         pve_vals=None, pve_threshold=None, K=None):
    # Single, model-agnostic PVE% array (RSS-reduction / SS_total, kinship-
    # aware — see compute_pve_all) reused for every sheet below instead of
    # recomputing a different (and previously inconsistent) r^2-based PVE
    # per model/per sheet.
    if pve_vals is None:
        pve_vals = compute_pve_all(y, G_filt, K=K)
    wb = Workbook()
    header_fill = PatternFill("solid", start_color="4A90D9", end_color="4A90D9")
    sig_fill    = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
    alt_fill    = PatternFill("solid", start_color="F8F9FA", end_color="F8F9FA")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    title_font  = Font(bold=True, color="2C5F8A", name="Calibri", size=13)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    bonf = 0.05 / len(list(all_results.values())[0][0])

    # Sheet 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum['A1'] = f"GWAS Complete Analysis Summary — Top {top_n} MTAs per Model"
    ws_sum['A1'].font = Font(bold=True, color="2C5F8A", name="Calibri", size=16)
    ws_sum.merge_cells('A1:H1')
    headers_sum = ["Model", "N_SNPs_Tested", "Bonferroni_Threshold",
                   "Suggestive_Threshold","Sig_SNPs_Bonferroni",
                   "Sig_SNPs_Suggestive", "Genomic_Lambda", "Min_Pvalue"]
    for col, h in enumerate(headers_sum, 1):
        cell = ws_sum.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        ws_sum.column_dimensions[get_column_letter(col)].width = 24
    for row_i, (mname, (pvals, betas, ses)) in enumerate(all_results.items(), 4):
        lam = np.median(stats.chi2.ppf(1-np.clip(pvals,1e-10,1),df=1)) / stats.chi2.ppf(0.5,df=1)
        n_snps = len(pvals)
        bonf_t = 0.05 / n_snps
        sug_t  = 1.0  / n_snps
        sig_bonf = int((pvals < bonf_t).sum())
        sig_sug  = int((pvals < sug_t).sum())
        row_data = [mname, n_snps, f"{bonf_t:.2e}", f"{sug_t:.2e}",
                    sig_bonf, sig_sug, round(float(lam),4), float(np.min(pvals))]
        fill = alt_fill if row_i%2==0 else PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
        for col, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=row_i, column=col, value=val)
            cell.fill = fill; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Threshold explanation sheet
    ws_exp = wb.create_sheet(title="Threshold_Explanation")
    ws_exp['A1'] = "Threshold Calculation Explanation"
    ws_exp['A1'].font = title_font
    ws_exp.merge_cells('A1:C1')
    explanation_rows = [
        ("Bonferroni Threshold", "Formula", "0.05 / m  where m = number of SNPs tested"),
        ("Bonferroni Threshold", "Rationale", "Controls genome-wide Type I error rate at 5%."),
        ("Bonferroni Threshold", "Assumption", "All SNP tests are independent (conservative)."),
        ("Suggestive Threshold", "Formula", "1 / m  where m = number of SNPs tested"),
        ("Suggestive Threshold", "Rationale", "Expects ~1 false positive per genome scan."),
        ("Suggestive Threshold", "Reference", "Lander & Kruglyak (1995) Nature Genetics."),
        ("OLS Note", "Correction", "No population structure correction."),
        ("MLM Note", "Correction", "Kinship matrix controls for relatedness."),
        ("EMMAX Note", "Correction", "Efficient Mixed Model Association with variance components."),
        ("FarmCPU Note", "Correction", "Iterative pseudo-QTN control reduces false positives."),
        ("GEMMA Note", "Correction", "Genomic control (GC correction) applied for inflation."),
    ]
    ws_exp.cell(row=3, column=1, value="Threshold").font = header_font
    ws_exp.cell(row=3, column=1).fill = header_fill
    ws_exp.cell(row=3, column=2, value="Aspect").font = header_font
    ws_exp.cell(row=3, column=2).fill = header_fill
    ws_exp.cell(row=3, column=3, value="Details").font = header_font
    ws_exp.cell(row=3, column=3).fill = header_fill
    ws_exp.column_dimensions['A'].width = 22
    ws_exp.column_dimensions['B'].width = 16
    ws_exp.column_dimensions['C'].width = 60
    for r_i, (t, a, d) in enumerate(explanation_rows, 4):
        ws_exp.cell(row=r_i, column=1, value=t).border = thin_border
        ws_exp.cell(row=r_i, column=2, value=a).border = thin_border
        ws_exp.cell(row=r_i, column=3, value=d).border = thin_border

    # Sheet per model: MTAs passing the permutation/LOD threshold
    for mname, (pvals, betas, ses) in all_results.items():
        sig_idx_sheet = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n,
                                             pve_vals=pve_vals, pve_threshold=pve_threshold)
        n_sig = len(sig_idx_sheet)
        sheet_label = f"MTA(LOD)" if lod_threshold is not None else f"Top{top_n}"
        ws = wb.create_sheet(title=f"{mname}_{sheet_label}"[:31])
        title_txt = (f"GWAS Results — {mname} Model — {n_sig} MTA(s) passing "
                    f"permutation/LOD threshold (LOD ≥ {lod_threshold:.2f})" if lod_threshold is not None
                    else f"GWAS Results — {mname} Model — Top {top_n} MTAs (no LOD threshold available)")
        ws['A1'] = title_txt
        ws['A1'].font = title_font
        ws.merge_cells('A1:K1')
        log_p = -np.log10(np.clip(pvals, 1e-300, 1))
        t_stats = np.where(ses > 0, betas / (ses + 1e-10), 0.0)
        reject_fdr, pvals_fdr, _, _ = multipletests(pvals, method='fdr_bh')
        reject_bonf, _, _, _ = multipletests(pvals, method='bonferroni')
        headers = ["Rank","SNP","Chr","Position","MAF","Beta","SE",
                   "T_statistic","P_value","neg_log10P","LOD","FDR_BH","Bonferroni_sig"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        col_widths = [8, 25, 8, 12, 10, 12, 12, 12, 14, 12, 10, 14, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        top_idx = sig_idx_sheet
        if len(top_idx) == 0:
            ws.cell(row=4, column=1,
                    value="No MTAs passed the permutation/LOD threshold for this model.")
        for rank, idx in enumerate(top_idx, 1):
            is_sig = bool(reject_bonf[idx]) or pvals[idx] < 5e-8
            fill = sig_fill if is_sig else (alt_fill if rank%2==0 else
                   PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF"))
            maf_v = maf_vals[idx] if idx < len(maf_vals) else np.nan
            row_data = [
                rank,
                str(snp_filt[idx]) if idx < len(snp_filt) else "",
                int(chr_filt[idx]) if idx < len(chr_filt) else "",
                int(pos_filt[idx]) if idx < len(pos_filt) else "",
                round(float(maf_v),4) if not np.isnan(maf_v) else "",
                round(float(betas[idx]),6),
                round(float(ses[idx]),6),
                round(float(t_stats[idx]),4),
                float(pvals[idx]),
                round(float(log_p[idx]),4),
                round(float(neglog10p_to_lod(log_p[idx])),3),
                round(float(pvals_fdr[idx]),6),
                "YES" if is_sig else "NO"
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=rank+3, column=col, value=val)
                cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if col == 9 and isinstance(val, float):
                    cell.number_format = '0.00E+00'

    # Sheet: Significant MTAs combined
    ws_mta = wb.create_sheet(title="Significant_MTAs")
    ws_mta['A1'] = (f"Significant MTAs — All Models — passing permutation/LOD threshold (LOD ≥ {lod_threshold:.2f})"
                   if lod_threshold is not None else f"Significant MTAs — All Models Combined — Top {top_n} per Model")
    ws_mta['A1'].font = title_font
    ws_mta.merge_cells('A1:N1')
    headers_mta = ["Model","Rank","SNP","Chr","Position","MAF","Beta","SE",
                   "T_stat","P_value","neg_log10P","LOD","FDR_BH_p","Bonferroni_sig","PVE_%"]
    for col, h in enumerate(headers_mta, 1):
        cell = ws_mta.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        ws_mta.column_dimensions[get_column_letter(col)].width = 16
    mta_row = 4
    for mname, (pvals, betas, ses) in all_results.items():
        log_p = -np.log10(np.clip(pvals, 1e-300, 1))
        t_stats = np.where(ses > 0, betas / (ses + 1e-10), 0.0)
        _, pvals_fdr, _, _ = multipletests(pvals, method='fdr_bh')
        reject_bonf, _, _, _ = multipletests(pvals, method='bonferroni')
        top_idx = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n)
        if len(top_idx) == 0:
            ws_mta.cell(row=mta_row, column=1, value=f"{mname}: no MTAs passed threshold")
            mta_row += 1
            continue
        for rank, idx in enumerate(top_idx, 1):
            pve_val = round(float(pve_vals[idx]), 4) if idx < len(pve_vals) else 0.0
            maf_v = maf_vals[idx] if idx < len(maf_vals) else np.nan
            row_data = [
                mname, rank,
                str(snp_filt[idx]) if idx < len(snp_filt) else "",
                int(chr_filt[idx]) if idx < len(chr_filt) else "",
                int(pos_filt[idx]) if idx < len(pos_filt) else "",
                round(float(maf_v),4) if not np.isnan(maf_v) else "",
                round(float(betas[idx]),6),
                round(float(ses[idx]),6),
                round(float(t_stats[idx]),4),
                float(pvals[idx]),
                round(float(log_p[idx]),4),
                round(float(neglog10p_to_lod(log_p[idx])),3),
                round(float(pvals_fdr[idx]),6),
                "YES" if bool(reject_bonf[idx]) else "NO",
                pve_val
            ]
            fill = sig_fill if mta_row%2==0 else PatternFill("solid",start_color="FFFDE7",end_color="FFFDE7")
            for col, val in enumerate(row_data, 1):
                cell = ws_mta.cell(row=mta_row, column=col, value=val)
                cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if col == 10 and isinstance(val, float):
                    cell.number_format = '0.00E+00'
            mta_row += 1

    # Sheet: QTN Effects
    ws_qtn = wb.create_sheet(title="QTN_Effects")
    ws_qtn['A1'] = (f"Quantitative Trait Nucleotides (QTNs) — passing permutation/LOD threshold (LOD ≥ {lod_threshold:.2f})"
                    if lod_threshold is not None else f"Quantitative Trait Nucleotides (QTNs) — Top {top_n} per Model")
    ws_qtn['A1'].font = title_font
    ws_qtn.merge_cells('A1:L1')
    headers_qtn = ["Model","Rank","SNP","Chr","Position","MAF",
                   "QTN_Effect_Beta","SE","95%_CI_Lower","95%_CI_Upper","P_value","LOD","PVE_%"]
    for col, h in enumerate(headers_qtn, 1):
        cell = ws_qtn.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        ws_qtn.column_dimensions[get_column_letter(col)].width = 17
    qtn_row = 4
    for mname, (pvals, betas, ses) in all_results.items():
        top_idx = get_significant_idx(pvals, lod_threshold, fallback_top_n=top_n)
        if len(top_idx) == 0:
            ws_qtn.cell(row=qtn_row, column=1, value=f"{mname}: no QTNs passed threshold")
            qtn_row += 1
            continue
        for rank, idx in enumerate(top_idx, 1):
            pve_val = round(float(pve_vals[idx]), 4) if idx < len(pve_vals) else 0.0
            maf_v = maf_vals[idx] if idx < len(maf_vals) else np.nan
            ci_lower = float(betas[idx]) - 1.96 * float(ses[idx])
            ci_upper = float(betas[idx]) + 1.96 * float(ses[idx])
            lod_val = round(float(neglog10p_to_lod(-np.log10(max(float(pvals[idx]), 1e-300)))), 3)
            fill = alt_fill if qtn_row%2==0 else PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
            row_data = [
                mname, rank,
                str(snp_filt[idx]) if idx < len(snp_filt) else "",
                int(chr_filt[idx]) if idx < len(chr_filt) else "",
                int(pos_filt[idx]) if idx < len(pos_filt) else "",
                round(float(maf_v),4) if not np.isnan(maf_v) else "",
                round(float(betas[idx]),6),
                round(float(ses[idx]),6),
                round(ci_lower,6), round(ci_upper,6),
                float(pvals[idx]), lod_val, pve_val
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws_qtn.cell(row=qtn_row, column=col, value=val)
                cell.fill = fill; cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if col == 11 and isinstance(val, float):
                    cell.number_format = '0.00E+00'
            qtn_row += 1

    # Sheet: SNP QC
    ws_qc = wb.create_sheet(title="SNP_QC")
    ws_qc['A1'] = "SNP Quality Control Statistics"
    ws_qc['A1'].font = title_font
    ws_qc.merge_cells('A1:F1')
    headers_qc = ["SNP","Chr","Position","MAF","Call_Rate","Category"]
    for col, h in enumerate(headers_qc, 1):
        cell = ws_qc.cell(row=3, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
        ws_qc.column_dimensions[get_column_letter(col)].width = 20
    call_rates = compute_call_rate(G_filt)
    for row_i in range(len(snp_filt)):
        maf_v = maf_vals[row_i] if row_i < len(maf_vals) else np.nan
        cr_v = call_rates[row_i] if row_i < len(call_rates) else np.nan
        cat = "Rare (<0.05)" if not np.isnan(maf_v) and maf_v < 0.05 else "Common"
        fill = alt_fill if row_i%2==0 else PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
        row_data = [
            str(snp_filt[row_i]),
            int(chr_filt[row_i]) if row_i < len(chr_filt) else "",
            int(pos_filt[row_i]) if row_i < len(pos_filt) else "",
            round(float(maf_v),4) if not np.isnan(maf_v) else "",
            round(float(cr_v),4) if not np.isnan(cr_v) else "",
            cat
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws_qc.cell(row=row_i+4, column=col, value=val)
            cell.fill = fill; cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    wb.save(output_path)
    return output_path


def save_all_plots_zip(plot_dict, out_dir, dpi=EXPORT_DPI):
    zip_path = os.path.join(out_dir, "gwas_plots.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, fig in plot_dict.items():
            if fig is not None:
                buf = io.BytesIO()
                fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                            facecolor=fig.get_facecolor())
                buf.seek(0)
                zf.writestr(f"{name}.png", buf.read())
    return zip_path


# ─────────────────────── MAIN PIPELINE ───────────────────────────────────────────
def run_gwas(file_obj, text_input, selected_models, maf_threshold, top_n, n_perm,
             lod_thr_manual, pve_thr,
             # Cofactor / advanced iterative-model settings
             iter_max_iter, farmcpu_bin_size, blink_ld_r2,
             mrmlm_screen_thresh, mrmlm_max_candidates, mrmlm_drop_thresh,
             mlm_n_pca, mrmlm_n_pca, fastmrmlm_n_pca, gemma_n_pca,
             # Manhattan cfg
             mh_w, mh_h, mh_fs, mh_fc, mh_ms, mh_alpha, mh_bg, mh_sig, mh_sug,
             mh_dot, mh_panel, mh_show_logp,
             # QQ cfg
             qq_w, qq_h, qq_fs, qq_fc, qq_ms, qq_dot, qq_line, qq_fill, qq_bg, qq_panel,
             # New Circos (density-ring) cfg
             nc_models, nc_theme, nc_rings_per_panel, nc_w, nc_h, nc_fs,
             # LD cfg
             ld_w, ld_h, ld_fs, ld_dot, ld_line,
             # PCA cfg (dark)
             pc_w, pc_h, pc_fs,
             # Kinship cfg
             ki_w, ki_h, ki_fs, ki_dot, ki_bg, ki_panel,
             # Effect cfg
             ef_w, ef_h, ef_fs, ef_pos, ef_neg, ef_fill, ef_bg,
             # Comparison cfg
             co_w, co_h, co_sig, co_sug,
             # Dashboard cfg
             db_w, db_h, db_dot):

    try:
        df, snp_cols, chroms, positions = parse_data(file_obj, text_input)
    except Exception as e:
        return [None] * 13 + [f"❌ Data parsing error: {e}"]

    y = df["Phenotype"].values.astype(float)
    ids = df["ID"].values
    G_raw = df[snp_cols].values.astype(float)
    G = impute_genotypes(G_raw)
    n, m = G.shape

    maf_vals_all = compute_maf(G)
    maf_thr = float(maf_threshold)
    keep = maf_vals_all >= maf_thr
    if keep.sum() < 2:
        keep = np.ones(m, dtype=bool)

    G_filt   = G[:, keep]
    snp_filt = np.array(snp_cols)[keep]
    chr_filt = chroms[keep] if len(chroms)==m else np.ones(G_filt.shape[1],dtype=int)
    pos_filt = positions[keep] if len(positions)==m else np.arange(G_filt.shape[1])
    maf_vals = maf_vals_all[keep]
    K = build_kinship(G_filt)

    model_funcs = {
        "OLS":        lambda: ols_gwas(y, G_filt),
        "MLM":        lambda: mlm_gwas(y, G_filt, K, n_pca=int(mlm_n_pca)),
        "EMMAX":      lambda: emmax_gwas(y, G_filt, K),
        "FarmCPU":    lambda: farmcpu_gwas(y, G_filt, K, chr_filt, pos_filt,
                                            max_iter=int(iter_max_iter),
                                            bin_size=int(farmcpu_bin_size)),
        "GEMMA":      lambda: gemma_gwas(y, G_filt, K,
                                          covariates=get_kinship_pcs(K, int(gemma_n_pca), n)),
        "BLINK":      lambda: blink_gwas(y, G_filt, K, chr_filt, pos_filt,
                                          max_iter=int(iter_max_iter),
                                          ld_r2=float(blink_ld_r2)),
        "mrMLM":      lambda: mrmlm_gwas(y, G_filt, K, chr_filt, pos_filt,
                                          screen_thresh=float(mrmlm_screen_thresh),
                                          max_candidates=int(mrmlm_max_candidates),
                                          drop_thresh=float(mrmlm_drop_thresh),
                                          n_pca=int(mrmlm_n_pca)),
        "FASTmrMLM":  lambda: fastmrmlm_gwas(y, G_filt, K, chr_filt, pos_filt,
                                              screen_thresh=float(mrmlm_screen_thresh),
                                              max_candidates=int(mrmlm_max_candidates),
                                              drop_thresh=float(mrmlm_drop_thresh),
                                              n_pca=int(fastmrmlm_n_pca)),
    }

    models_to_run = [mdl for mdl in (selected_models or list(model_funcs.keys()))
                     if mdl in model_funcs]
    if not models_to_run:
        models_to_run = list(model_funcs.keys())

    all_results = {}
    primary_p = primary_b = primary_se = None

    for mname in models_to_run:
        try:
            p, b, se = model_funcs[mname]()
            all_results[mname] = (p, b, se)
            if primary_p is None:
                primary_p, primary_b, primary_se = p, b, se
        except Exception as ex:
            print(f"Model {mname} failed: {ex}")

    if primary_p is None:
        return [None] * 13 + ["❌ No models succeeded", None]

    # ── Permutation (LOD) threshold — THE primary significance criterion ────
    # Churchill & Doerge-style empirical threshold from this panel, expressed
    # on the LOD scale. Everything downstream (Manhattan/Effect-size
    # highlighting, Excel MTA sheets) uses this to decide what actually
    # counts as an MTA, instead of a blind top-N slice.
    try:
        n_perm_eff = min(int(n_perm) if n_perm else 100, 200)
        lod_threshold, _, _ = permutation_threshold_lod(
            y, G_filt, K, n_perm=n_perm_eff)
    except Exception as ex:
        print(f"Permutation threshold failed: {ex}")
        lod_threshold = None
        n_perm_eff = 0

    # Manual LOD threshold override (0 = auto/permutation-derived, the default)
    lod_manual = float(lod_thr_manual) if lod_thr_manual else 0.0
    lod_threshold_auto = lod_threshold
    if lod_manual > 0:
        lod_threshold = lod_manual

    # PVE% (RSS-reduction / SS_total, kinship-aware) for EVERY marker.
    # Always computed now -- it's used both for the optional PVE% filter
    # threshold below AND as the PVE_% column in the Excel export, so both
    # need to reference the same array rather than each computing their own
    # (previously inconsistent) version.
    pve_threshold = float(pve_thr) if pve_thr else 0.0
    pve_vals_all = compute_pve_all(y, G_filt, K=K)

    top_n = int(top_n)

    # ── Build per-plot configs ────────────────────────────────────────────────
    # All build_cfg calls now correctly pass 'fc' as the 4th positional arg
    cfg_mh = build_cfg(mh_w, mh_h, mh_fs, mh_fc,
                       bg=mh_bg, panel=mh_panel,
                       sig_col=mh_sig, sug_col=mh_sug,
                       marker_size=mh_ms, alpha=mh_alpha,
                       dot_col=mh_dot)

    cfg_qq = build_cfg(qq_w, qq_h, qq_fs, qq_fc,
                       bg=qq_bg, panel=qq_panel,
                       marker_size=qq_ms,
                       dot_col=qq_dot, line_col=qq_line, fill_col=qq_fill)

    cfg_ld = build_cfg(ld_w, ld_h, ld_fs, "#2D3142",
                       dot_col=ld_dot, line_col=ld_line)

    # PCA — light theme (matches every other plot; no dark background)
    cfg_pc = build_cfg(pc_w, pc_h, pc_fs, TEXT_DARK,
                       gc=GRID_LIGHT, bg=BG_LIGHT, panel=PANEL_LIGHT,
                       sig_col=ACCENT_TEAL_D, sug_col=ACCENT_GOLD_D,
                       marker_size=55, alpha=0.80)

    cfg_ki = build_cfg(ki_w, ki_h, ki_fs, "#2D3142",
                       bg=ki_bg, panel=ki_panel, dot_col=ki_dot)

    cfg_ef = build_cfg(ef_w, ef_h, ef_fs, "#2D3142",
                       bg=ef_bg,
                       dot_col=ef_pos, line_col=ef_neg, fill_col=ef_fill)

    cfg_co = build_cfg(co_w, co_h, 7, "#2D3142",
                       sig_col=co_sig, sug_col=co_sug)

    cfg_db = build_cfg(db_w, db_h, 9, "#2D3142",
                       dot_col=db_dot)

    # Generate all plots
    plot_dict = {}
    plot_dict["01_Manhattan_Primary"] = plot_manhattan(
        primary_p, chr_filt, pos_filt,
        "Manhattan Plot", models_to_run[0], cfg_mh, top_n, lod_threshold=lod_threshold,
        pve_vals=pve_vals_all, pve_threshold=pve_threshold, show_logp_axis=bool(mh_show_logp))
    plot_dict["02_QQ_Primary"] = plot_qq_single(
        primary_p, "QQ Plot", models_to_run[0],
        MODEL_COLORS.get(models_to_run[0], "#A78BFA"), cfg_qq)
    plot_dict["03_QQ_All_Models"] = plot_qq_all_models(all_results, cfg_qq)
    plot_dict["04_QQ_Overlay"]    = plot_qq_overlay(all_results, cfg_qq)
    plot_dict["06_LD_Decay"]      = plot_ld_decay(G_filt, pos_filt, cfg_ld)
    plot_dict["07_PCA"]           = plot_pca(G_filt, y, ids, cfg_pc)
    plot_dict["08_Kinship"]       = plot_kinship(K, ids, cfg_ki)
    plot_dict["09_Effect_Sizes"]  = plot_effect_sizes(primary_p, primary_b, snp_filt, cfg_ef, top_n, lod_threshold=lod_threshold,
                                                        pve_vals=pve_vals_all, pve_threshold=pve_threshold)
    plot_dict["10_All_Models_MH"] = plot_model_comparison_manhattan(
        all_results, chr_filt, pos_filt, cfg_co, top_n, lod_threshold=lod_threshold)
    plot_dict["11_Dashboard"]     = plot_summary_dashboard(y, maf_vals, all_results, G_filt, cfg_db)

    nc_wanted = [m for m in (nc_models or []) if m in all_results]
    if not nc_wanted:
        nc_wanted = [m for m in models_to_run if m in all_results]
    cfg_nc = {"width": nc_w, "height": nc_h, "font_size": nc_fs}
    plot_dict["12_Circos_Density"] = plot_circos_density(
        all_results, chr_filt, pos_filt, nc_wanted, theme_name=nc_theme,
        rings_per_panel=int(nc_rings_per_panel), cfg=cfg_nc, lod_threshold=lod_threshold)

    def _pil(fig):
        return fig_to_pil(fig) if fig is not None else None

    imgs = [_pil(plot_dict[k]) for k in sorted(plot_dict.keys())]

    # Save files
    tmp_dir = tempfile.mkdtemp()
    zip_path = save_all_plots_zip(plot_dict, tmp_dir)
    excel_path = os.path.join(tmp_dir, "gwas_results_top15.xlsx")
    export_results_excel(all_results, snp_filt, chr_filt, pos_filt,
                         y, G_filt, maf_vals, excel_path, top_n, lod_threshold=lod_threshold,
                         pve_vals=pve_vals_all, pve_threshold=pve_threshold)
    plt.close('all')

    # Build threshold explanation string
    thresh_explanation = ""
    for mname, (p, _, _) in all_results.items():
        _, _, expl = compute_thresholds(len(p), mname)
        thresh_explanation += expl + "\n"

    bonf = 0.05 / len(primary_p)
    lambdas_str = "\n  ".join([
        f"{mdl}: λ={np.median(stats.chi2.ppf(1-np.clip(p,1e-10,1),df=1))/stats.chi2.ppf(0.5,df=1):.4f}  "
        f"sig(Bonf)={int((p<0.05/len(p)).sum())}  sig(Sug)={int((p<1.0/len(p)).sum())}"
        for mdl, (p, _, _) in all_results.items()
    ])
    lam = (np.median(stats.chi2.ppf(1-np.clip(primary_p,1e-10,1),df=1)) /
           stats.chi2.ppf(0.5, df=1))

    n_sig_lod_primary = len(get_significant_idx(primary_p, lod_threshold, fallback_top_n=top_n,
                                                 pve_vals=pve_vals_all, pve_threshold=pve_threshold))
    if lod_manual > 0:
        lod_note = f"{lod_threshold:.2f}  (MANUAL override — auto/permutation value was {lod_threshold_auto if lod_threshold_auto is not None else 'unavailable'})"
    elif lod_threshold is not None:
        lod_note = f"{lod_threshold:.2f}  ({n_perm_eff} permutations, α=0.05) — PRIMARY significance criterion"
    else:
        lod_note = "failed to compute — falling back to top-N display"
    pve_note = f"≥{pve_threshold:.1f}% (applied jointly with LOD threshold)" if pve_threshold > 0 else "off"

    summary = (
        f"{'='*68}\n"
        f"  GWAS ANALYSIS SUMMARY\n"
        f"{'='*68}\n"
        f"  Samples: {n}  |  SNPs (raw): {m}  |  SNPs (filtered): {G_filt.shape[1]}\n"
        f"  MAF threshold: ≥{maf_thr:.2f}  |  Models: {', '.join(all_results.keys())}\n"
        f"  Permutation / LOD threshold: {lod_note}\n"
        f"  PVE threshold: {pve_note}\n"
        f"  MTAs identified (primary model, {models_to_run[0]}): {n_sig_lod_primary}"
        f"{' (passing LOD threshold' + (' & PVE filter)' if pve_threshold > 0 else ')') if lod_threshold is not None else ' (fallback top-N, no threshold available)'}\n"
        f"{'─'*68}\n"
        f"  PER-MODEL RESULTS:\n  {lambdas_str}\n"
        f"{'─'*68}\n"
        f"  Primary model ({models_to_run[0]}):\n"
        f"    Bonferroni threshold : {bonf:.2e}  (-log₁₀ = {-np.log10(bonf):.2f}, LOD = {neglog10p_to_lod(-np.log10(bonf)):.2f})\n"
        f"    Suggestive threshold : {1/len(primary_p):.2e}  (-log₁₀ = {-np.log10(1/len(primary_p)):.2f}, LOD = {neglog10p_to_lod(-np.log10(1/len(primary_p))):.2f})\n"
        f"    Significant (Bonf.)  : {int((primary_p<bonf).sum())}\n"
        f"    Significant (Sug.)   : {int((primary_p<1/len(primary_p)).sum())}\n"
        f"    Genomic Inflation λ  : {lam:.4f}\n"
        f"{'─'*68}\n"
        f"  THRESHOLD EXPLANATIONS (see Excel sheet 'Threshold_Explanation'):\n\n"
        + thresh_explanation +
        f"{'─'*68}\n"
        f"  Excel : gwas_results_top15.xlsx\n"
        f"  Plots : gwas_plots.zip ({len(plot_dict)} PNG files)\n"
        f"{'='*68}\n"
    )

    summary_path = os.path.join(tmp_dir, "gwas_summary.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary)

    cache_state = {
        "all_results": all_results,       # {model: (pvals, betas, ses)}
        "snp_filt": snp_filt,
        "chr_filt": chr_filt,
        "pos_filt": pos_filt,
        "models_to_run": models_to_run,
        "primary_model": models_to_run[0],
        "primary_p": primary_p,
        "top_n": top_n,
        "lod_threshold": lod_threshold,
        "pve_vals": pve_vals_all,
        "pve_threshold": pve_threshold,
        "show_logp_axis": bool(mh_show_logp),
    }

    return imgs[:11] + [excel_path, zip_path, summary_path, summary, cache_state]


# ─────────────────────── LIVE MANHATTAN RECOLOR ───────────────────────────────────
def refresh_manhattan_live(state, mh_w, mh_h, mh_fs, mh_fc, mh_ms, mh_alpha,
                            mh_bg, mh_sig, mh_sug, mh_dot, mh_panel, mh_show_logp):
    """
    Re-renders ONLY the primary Manhattan plot from the cached p-values of the
    last completed run — no re-fitting of any model. Bound to every Manhattan color
    picker / slider's .change() event so tweaking a color updates the image
    instantly instead of waiting for the whole 🚀 Run pipeline to repeat.
    """
    if not state or "primary_p" not in state:
        return None
    cfg = build_cfg(mh_w, mh_h, mh_fs, mh_fc, bg=mh_bg, panel=mh_panel,
                     sig_col=mh_sig, sug_col=mh_sug, marker_size=mh_ms,
                     alpha=mh_alpha, dot_col=mh_dot)
    fig = plot_manhattan(state["primary_p"], state["chr_filt"], state["pos_filt"],
                          "Manhattan Plot", state["primary_model"], cfg, state["top_n"],
                          lod_threshold=state.get("lod_threshold"),
                          pve_vals=state.get("pve_vals"), pve_threshold=state.get("pve_threshold"),
                          show_logp_axis=bool(mh_show_logp))
    img = fig_to_pil(fig)
    plt.close(fig)
    return img

# ─────────────────────── COMPARE MODELS — COMMON MTA ──────────────────────────────
def _sig_index_set(pvals, mode, top_n, lod_threshold=None):
    """Return the set of marker indices considered 'significant' for one
    model, under Bonferroni, a fixed top-N rank cutoff, or (recommended)
    the panel's permutation/LOD threshold -- the same primary criterion
    used throughout the rest of the pipeline."""
    m = len(pvals)
    if mode.startswith("LOD"):
        idx = get_significant_idx(np.asarray(pvals), lod_threshold, fallback_top_n=int(top_n))
        return set(idx.tolist())
    elif mode.startswith("Bonferroni"):
        bonf = 0.05 / m
        idx = np.where(pvals < bonf)[0]
    else:
        idx = np.argsort(pvals)[:int(top_n)]
    return set(idx.tolist())

COMPARE_COLORS = ["#3B5BDB", "#14B8A6", "#DB4C8C", "#5FA624", "#B054E0",
                   "#E0972E", "#2AB0C5", "#D65B3E"]

def plot_common_mta(state, chosen_models, mode, top_n, cfg=None):
    """
    Two-panel comparison figure:
      Top    — bar chart: how many markers are shared by exactly k of the
               chosen models (the 'consensus degree' distribution). The bar
               for markers shared by ALL chosen models is highlighted gold.
      Bottom — consensus Manhattan: every chosen model's -log10(p) track
               overlaid in its own color (thin, semi-transparent), with
               markers significant in EVERY chosen model circled as gold
               diamonds so you can see where the models agree at a glance.
    """
    if not state or "all_results" not in state or not chosen_models:
        return None, pd.DataFrame()

    all_results = state["all_results"]
    chroms = np.asarray(state["chr_filt"])
    positions = np.asarray(state["pos_filt"])
    snp_names = np.asarray(state["snp_filt"])
    lod_threshold = state.get("lod_threshold")

    sig_sets = {}
    for mname in chosen_models:
        if mname not in all_results:
            continue
        p, _, _ = all_results[mname]
        sig_sets[mname] = _sig_index_set(np.asarray(p), mode, top_n, lod_threshold)

    if not sig_sets:
        return None, pd.DataFrame()

    n_models = len(sig_sets)
    union_idx = sorted(set().union(*sig_sets.values()))
    degree = {i: sum(1 for s in sig_sets.values() if i in s) for i in union_idx}
    consensus_idx = [i for i, d in degree.items() if d == n_models]

    if cfg is None:
        cfg = build_cfg(16, 10, 10, "#20233D")

    fig = plt.figure(figsize=(cfg["width"], cfg["height"]))
    fig.patch.set_facecolor(cfg["bg_color"])
    gs = gridspec.GridSpec(2, 1, height_ratios=[1, 1.6], hspace=0.5)

    # ── Panel 1: consensus-degree bar chart ────────────────────────────
    ax1 = fig.add_subplot(gs[0])
    ax1.set_facecolor(cfg["panel_color"])
    deg_counts = [sum(1 for d in degree.values() if d == k) for k in range(1, n_models + 1)]
    bar_colors = ["#B9C2E8"] * (n_models - 1) + ["#F59E0B"] if n_models > 0 else []
    ax1.bar(range(1, n_models + 1), deg_counts, color=bar_colors, edgecolor="#33333340")
    ax1.set_xticks(range(1, n_models + 1))
    ax1.set_xticklabels([f"{k}/{n_models}\nmodel(s)" for k in range(1, n_models + 1)],
                        fontsize=cfg["font_size"] - 1, color=cfg["font_color"])
    ax1.set_ylabel("# MTAs", color=cfg["font_color"], fontsize=cfg["font_size"])
    ax1.set_title(f"Consensus across {n_models} selected model(s) — "
                   f"{len(consensus_idx)} MTA(s) found by ALL of them",
                   color=cfg["font_color"], fontsize=cfg["font_size"] + 1)
    ax1.tick_params(colors=cfg["font_color"])
    for spine in ax1.spines.values():
        spine.set_edgecolor("#CCCCCC")
    ax1.grid(axis="y", color="#E8ECF4", linewidth=0.6, linestyle="--", alpha=0.7)

    # ── Panel 2: overlaid consensus Manhattan ──────────────────────────
    ax2 = fig.add_subplot(gs[1])
    ax2.set_facecolor(cfg["panel_color"])
    chrom_list = sorted(set(chroms.tolist()))
    gap = max(positions) * 0.018 + 1
    x_offset = 0
    x_coords = np.zeros(len(chroms))
    tick_pos, tick_lab = [], []
    for ch in chrom_list:
        mask = chroms == ch
        pos = positions[mask]
        if len(pos) == 0:
            continue
        order = np.argsort(pos)
        x_vals = pos[order] - pos.min() + x_offset
        x_coords[np.where(mask)[0][order]] = x_vals
        tick_pos.append(x_vals.mean())
        tick_lab.append(f"Chr{ch}")
        x_offset += (pos.max() - pos.min()) + gap

    for i, mname in enumerate(sig_sets.keys()):
        p, _, _ = all_results[mname]
        log_p = -np.log10(np.clip(np.asarray(p), 1e-300, 1))
        color = MODEL_COLORS.get(mname, COMPARE_COLORS[i % len(COMPARE_COLORS)])
        ax2.scatter(x_coords, log_p, s=8, alpha=0.35, color=color, linewidths=0,
                    label=mname)

    if consensus_idx:
        ax2.scatter(x_coords[consensus_idx],
                    [max(-np.log10(max(all_results[m][0][i], 1e-300)) for m in sig_sets) for i in consensus_idx],
                    marker="D", s=90, facecolor=ACCENT_GOLD, edgecolor="#333333",
                    linewidth=1.0, zorder=10, label=f"Common to all {n_models}")

    ax2.set_xticks(tick_pos)
    ax2.set_xticklabels(tick_lab, rotation=45, fontsize=cfg["font_size"] - 1,
                        color=cfg["font_color"])
    ax2.set_ylabel("-log₁₀(p)", color=cfg["font_color"], fontsize=cfg["font_size"])
    ax2.set_xlabel("Chromosome", color=cfg["font_color"], fontsize=cfg["font_size"])
    ax2.tick_params(colors=cfg["font_color"])
    for spine in ax2.spines.values():
        spine.set_edgecolor("#CCCCCC")
    ax2.grid(axis="y", color="#E8ECF4", linewidth=0.6, linestyle="--", alpha=0.7)
    ax2.legend(fontsize=cfg["font_size"] - 2, framealpha=0.9, facecolor=cfg["bg_color"],
              edgecolor="#CCCCCC", loc="upper right", ncol=2)

    fig.tight_layout()

    # ── Table of shared / unique MTAs ──────────────────────────────────
    rows = []
    for i in sorted(union_idx, key=lambda i: (-degree[i], min(all_results[m][0][i] for m in sig_sets if i in sig_sets[m]))):
        row = {
            "SNP": snp_names[i], "Chr": int(chroms[i]), "Pos": int(positions[i]),
            "N_Models_Sig": degree[i],
            "Models": ", ".join(m for m in sig_sets if i in sig_sets[m]),
        }
        for m in sig_sets:
            p, _, _ = all_results[m]
            pv = float(p[i])
            row[f"{m}_p"] = f"{pv:.2e}"
            row[f"{m}_LOD"] = f"{float(neglog10p_to_lod(-np.log10(max(pv, 1e-300)))):.2f}"
        rows.append(row)
    df = pd.DataFrame(rows)

    return fig, df


def run_compare_models(state, chosen_models, mode, top_n, cmp_w, cmp_h, cmp_fs):
    cfg = build_cfg(cmp_w, cmp_h, cmp_fs, "#20233D")
    fig, df = plot_common_mta(state, chosen_models, mode, top_n, cfg)
    if fig is None:
        return None, pd.DataFrame({"Message": ["Run the pipeline first, then pick ≥1 model here."]}), None, None
    img = fig_to_pil(fig)

    tmp_dir = tempfile.mkdtemp(prefix="gwas_common_mta_")
    fig_path = os.path.join(tmp_dir, "common_MTA_figure.png")
    fig.savefig(fig_path, format="png", dpi=EXPORT_DPI, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)

    table_path = os.path.join(tmp_dir, "common_MTA_table.xlsx")
    with pd.ExcelWriter(table_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Common_MTAs", index=False)
        ws = writer.sheets["Common_MTAs"]
        header_fill = PatternFill("solid", start_color="4A90D9", end_color="4A90D9")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value is not None else 0 for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 40)

    return img, df, fig_path, table_path



SAMPLE_DATA = """ID,Phenotype,chr1_134521_SNP1,chr1_134810_SNP2,chr1_178766_SNP3,chr1_210328_SNP4
chromosome,1,1,1,1
IRRI-1,25.34,0,0,0,2
IRRI-75,23.34,0,0,0,0
IRRI-84,26.67,2,0,2,0
IRRI-93,23,0,0,0,1
IRRI-9,28.34,0,0,0,0
IRRI-17,18.67,0,0,0,0
IRRI-25,20,2,0,2,0
IRRI-34,37,0,0,0,0
IRRI-42,30.67,0,0,0,0
IRRI-50,22,0,0,0,2"""
# =============================================================================
# SHINY FOR PYTHON USER INTERFACE
# =============================================================================
# The GWAS/statistical engine above is intentionally unchanged.  Only the
# Gradio presentation/event layer has been replaced with Shiny for Python.

from types import SimpleNamespace
from pathlib import Path
from shiny import App, Inputs, Outputs, Session, reactive, render, ui, req

SHINY_CSS = """
:root { --ink:#1A1D29; --ink-soft:#565D75; --line:#E1E4EC; --paper:#FFFFFF;
        --canvas:#F7F8FB; --accent:#2B4C7E; }
body { background:var(--canvas); color:var(--ink); font-family:Inter,Arial,sans-serif; }
.gwas-header { background:var(--paper); border:1px solid var(--line);
  border-bottom:3px solid var(--accent); border-radius:6px; padding:20px 26px; margin-bottom:18px; }
.gwas-header h1 { margin:0 0 4px 0; font-size:25px; font-weight:700; }
.gwas-header p { color:var(--ink-soft); font-size:13px; margin:3px 0; }
.section-title { font-size:12px; font-weight:700; text-transform:uppercase;
  letter-spacing:.5px; padding-bottom:8px; margin-bottom:12px; border-bottom:1px solid var(--line); }
.info-box { background:#EEF2F8; border:1px solid var(--line); border-left:3px solid var(--accent);
  border-radius:4px; padding:10px 14px; margin:8px 0; font-size:12px; line-height:1.5; }
.card { background:var(--paper); border:1px solid var(--line); border-radius:6px; padding:18px; margin-bottom:14px; }
.model-chip { display:inline-block; padding:3px 9px; border-radius:4px; color:#fff; font-size:11px; font-weight:600; margin:2px; }
.chip-OLS{background:#3E7BD6}.chip-MLM{background:#14B8A6}.chip-EMMAX{background:#DB4C8C}
.chip-FarmCPU{background:#5FA624}.chip-GEMMA{background:#B054E0}.chip-BLINK{background:#E0972E}
.chip-mrMLM{background:#2AB0C5}.chip-FASTmrMLM{background:#D65B3E}
.summary-box { background:#F7F8FB; border:1px solid var(--line); border-radius:5px;
  padding:14px; white-space:pre-wrap; font-family:monospace; font-size:12px; max-height:760px; overflow:auto; }
.plot-box { background:#fff; border:1px solid var(--line); border-radius:6px; padding:10px; margin-bottom:14px; }
.shiny-output-error { color:#A00000; }
button.btn-primary { background:#2B4C7E; border-color:#2B4C7E; }
"""

# Helpers to make the UI readable.
def input_color(id, label, value):
    # Shiny core has no dedicated color-picker input; HTML color input is used
    # as a normal Shiny text-compatible value through a small custom input.
    # The companion text field is the reliable server-side value.
    return ui.input_text(id, label, value=value, placeholder="#RRGGBB")

def slider(id, label, lo, hi, value, step):
    return ui.input_slider(id, label, min=lo, max=hi, value=value, step=step)

def card(*children):
    return ui.div(*children, class_="card")

def chips():
    return ui.HTML("".join(f'<span class="model-chip chip-{m}">{m}</span>' for m in ALL_MODELS))

app_ui = ui.page_fluid(
    ui.tags.style(SHINY_CSS),
    ui.div(
        ui.h1("GWAS Analysis Pipeline"),
        ui.p("OLS · MLM · EMMAX · FarmCPU · GEMMA · BLINK · mrMLM · FASTmrMLM"),
        ui.p("Bonferroni and suggestive thresholds on every plot · live Manhattan recolor · cross-model MTA comparison"),
        class_="gwas-header",
    ),
    ui.navset_tab(
        ui.nav_panel(
            "1. Data & Models",
            ui.layout_columns(
                ui.column(4,
                    card(
                        ui.div("Data Input", class_="section-title"),
                        ui.input_file("file_in", "Upload CSV / TSV / TXT", accept=[".csv", ".tsv", ".txt"]),
                        ui.input_text_area("text_in", "Or paste CSV data", value=SAMPLE_DATA, rows=8),
                    ),
                    card(
                        ui.div("Model Selection", class_="section-title"), chips(),
                        ui.p("Colors match every plot legend and the Compare Models tab."),
                        ui.input_checkbox_group("model_sel", "Select models to run",
                            choices=ALL_MODELS,
                            selected=["OLS", "MLM", "EMMAX", "FarmCPU", "GEMMA"]),
                        slider("maf_thr", "MAF Threshold", 0, .5, .05, .01),
                        slider("top_n_sl", "Fallback Top-N", 5, 50, 15, 1),
                        slider("n_perm", "Permutations for LOD threshold", 20, 200, 100, 10),
                        ui.div("Markers are highlighted as MTAs when they pass the LOD threshold. More permutations give a more stable threshold but take longer.", class_="info-box"),
                        ui.div("Significance & MTA Selection", class_="section-title"),
                        slider("lod_thr_manual", "Manual LOD threshold (0 = automatic)", 0, 10, 0, .1),
                        slider("pve_thr", "Minimum PVE % (0 = off)", 0, 100, 0, 1),
                        ui.input_action_button("run_btn", "Run Complete GWAS Pipeline", class_="btn-primary"),
                    ),
                    card(
                        ui.div("Downloads", class_="section-title"),
                        ui.download_button("excel_dl", "Download Excel results"),
                        ui.download_button("zip_dl", "Download all plots (ZIP)"),
                        ui.download_button("summary_dl", "Download summary (TXT)"),
                    ),
                ),
                ui.column(8,
                    card(
                        ui.div("Analysis Summary + Threshold Explanations", class_="section-title"),
                        ui.output_text_verbatim("summary_out"),
                    )
                ),
                col_widths=[4,8], gap="20px",
            ),
        ),
        ui.nav_panel(
            "2. Cofactors / Advanced",
            ui.div("Left group affects multi-locus models (FarmCPU, BLINK, mrMLM, FASTmrMLM). Kinship-PC covariates are used by MLM, mrMLM, FASTmrMLM and GEMMA when selected.", class_="info-box"),
            ui.layout_columns(
                card(
                    ui.div("FarmCPU / BLINK Pseudo-QTN Iteration", class_="section-title"),
                    slider("iter_max_iter", "Max FEM/REM iterations", 2, 15, 8, 1),
                    slider("farmcpu_bin_size", "FarmCPU bin size (bp)", 100000, 5000000, 1000000, 100000),
                    slider("blink_ld_r2", "BLINK LD r² pruning cutoff", .1, .99, .7, .01),
                ),
                card(
                    ui.div("mrMLM / FASTmrMLM Two-Stage Screening", class_="section-title"),
                    slider("mrmlm_screen_thresh", "Stage-1 screening p-value", .001, .05, .01, .001),
                    slider("mrmlm_max_candidates", "Max candidate QTNs", 5, 40, 20, 1),
                    slider("mrmlm_drop_thresh", "Stage-2 drop threshold", .01, .2, .05, .01),
                ),
                col_widths=[6,6], gap="20px",
            ),
            card(
                ui.div("Kinship-PC Covariates (per model)", class_="section-title"),
                ui.p('Number of top kinship eigenvectors used as fixed “Q” structure covariates. Set to 0 to disable.'),
                ui.layout_columns(
                    slider("mlm_n_pca", "MLM — kinship PCs", 0, 10, 3, 1),
                    slider("gemma_n_pca", "GEMMA — kinship PCs", 0, 10, 0, 1),
                    slider("mrmlm_n_pca", "mrMLM — kinship PCs", 0, 10, 3, 1),
                    slider("fastmrmlm_n_pca", "FASTmrMLM — kinship PCs", 0, 10, 3, 1),
                    col_widths=[3,3,3,3], gap="15px",
                ),
            ),
        ),
        ui.nav_panel(
            "3. Plot Customization",
            ui.div("Manhattan updates live after a completed GWAS run; changing these settings does not re-fit models.", class_="info-box"),
            ui.accordion(
                ui.accordion_panel("Manhattan Plot Options",
                    ui.input_checkbox("mh_show_logp", "Show -log₁₀(p) on secondary right axis", True),
                    ui.layout_columns(
                        slider("mh_w", "Width", 8,24,16,1), slider("mh_h", "Height",3,12,5,.5), slider("mh_fs", "Font Size",6,18,9,1),
                        input_color("mh_fc", "Font Color", "#2D3142"), input_color("mh_sig", "Bonferroni Line Color", "#E15759"), input_color("mh_sug", "Suggestive Line Color", "#4E79A7"),
                        slider("mh_ms", "Marker Size",2,60,12,1), slider("mh_alpha", "Marker Alpha",.1,1,.75,.05),
                        input_color("mh_bg", "Background Color", "#FFFFFF"), input_color("mh_panel", "Panel Color", "#FAFBFF"), input_color("mh_dot", "Dot Color", "#4E79A7"),
                        col_widths=[4,4,4], gap="15px",
                    ), open=True,
                ),
                ui.accordion_panel("QQ Plot Options",
                    ui.layout_columns(
                        slider("qq_w", "Width",5,16,7,.5), slider("qq_h", "Height",5,16,7,.5), slider("qq_fs", "Font Size",6,18,10,1),
                        input_color("qq_fc", "Font Color", "#2D3142"), slider("qq_ms", "Marker Size",5,80,22,1),
                        input_color("qq_dot", "Dot Color", "#A78BFA"), input_color("qq_line", "Regression Line Color", "#E15759"), input_color("qq_fill", "CI Fill Color", "#A78BFA"),
                        input_color("qq_bg", "Background Color", "#FFFFFF"), input_color("qq_panel", "Panel Color", "#FAFBFF"),
                        col_widths=[4,4,4], gap="15px",
                    ),
                ),
                ui.accordion_panel("Circos Plot Options (Density Ring)",
                    ui.div("Outer band = SNP density heatmap per chromosome. Inner rings show selected model results.", class_="info-box"),
                    ui.input_checkbox_group("nc_models", "Models to show", choices=ALL_MODELS, selected=["OLS","MLM","EMMAX","FarmCPU","GEMMA"]),
                    ui.input_radio_buttons("nc_theme", "Color Theme", choices=["Bright","Dark"], selected="Bright", inline=True),
                    ui.layout_columns(
                        slider("nc_rings_per_panel", "Rings per panel",2,4,3,1), slider("nc_w", "Width",10,30,20,1), slider("nc_h", "Height",10,30,20,1), slider("nc_fs", "Font Size",6,22,11,1),
                        col_widths=[3,3,3,3], gap="15px"),
                ),
                ui.accordion_panel("LD Decay Options",
                    ui.layout_columns(slider("ld_w","Width",8,24,16,1),slider("ld_h","Height",3,12,6,.5),slider("ld_fs","Font Size",6,18,9,1),input_color("ld_dot","Dot Color","#4E79A7"),input_color("ld_line","Trend Line Color","#E15759"),col_widths=[3,3,3,3,3],gap="15px")),
                ui.accordion_panel("PCA Options",
                    ui.layout_columns(slider("pc_w","Width",8,22,14,1),slider("pc_h","Height",4,14,10,1),slider("pc_fs","Font Size",6,18,9,1),col_widths=[4,4,4],gap="15px")),
                ui.accordion_panel("Kinship / Dendrogram Options",
                    ui.layout_columns(slider("ki_w","Width",8,22,14,1),slider("ki_h","Height",6,20,12,1),slider("ki_fs","Font Size",6,18,9,1),input_color("ki_dot","Dot/Bar Color","#4E79A7"),input_color("ki_bg","Background Color","#FFFFFF"),input_color("ki_panel","Panel Color","#FAFBFF"),col_widths=[4,4,4],gap="15px")),
                ui.accordion_panel("Effect Sizes Options",
                    ui.layout_columns(slider("ef_w","Width",8,22,14,1),slider("ef_h","Height",5,18,10,1),slider("ef_fs","Font Size",6,18,9,1),input_color("ef_pos","Positive Effect Color","#4E79A7"),input_color("ef_neg","Negative Effect Color","#E15759"),input_color("ef_fill","Distribution Fill Color","#A78BFA"),input_color("ef_bg","Background Color","#FFFFFF"),col_widths=[4,4,4],gap="15px")),
                ui.accordion_panel("All-Models Manhattan Options",
                    ui.layout_columns(slider("co_w","Width",8,24,16,1),slider("co_h","Height per model",2,8,3.5,.5),input_color("co_sig","Bonferroni Line Color","#E15759"),input_color("co_sug","Suggestive Line Color","#4E79A7"),col_widths=[3,3,3,3],gap="15px")),
                ui.accordion_panel("Dashboard Options",
                    ui.layout_columns(slider("db_w","Width",8,24,15,1),slider("db_h","Height",5,18,10,1),input_color("db_dot","Bar/Dot Color","#4E79A7"),col_widths=[4,4,4],gap="15px")),
            ),
        ),
        ui.nav_panel(
            "4. Results",
            ui.navset_tab(
                ui.nav_panel("Manhattan", ui.output_image("img1", width="100%", height="650px")),
                ui.nav_panel("QQ Primary", ui.output_image("img2", width="100%", height="650px")),
                ui.nav_panel("QQ All Models", ui.output_image("img3", width="100%", height="650px")),
                ui.nav_panel("QQ Overlay", ui.output_image("img4", width="100%", height="650px")),
                ui.nav_panel("LD Decay", ui.output_image("img5", width="100%", height="650px")),
                ui.nav_panel("PCA", ui.output_image("img6", width="100%", height="850px")),
                ui.nav_panel("Kinship", ui.output_image("img7", width="100%", height="850px")),
                ui.nav_panel("Effect Sizes", ui.output_image("img8", width="100%", height="850px")),
                ui.nav_panel("All Models MH", ui.output_image("img9", width="100%", height="650px")),
                ui.nav_panel("Dashboard", ui.output_image("img10", width="100%", height="850px")),
                ui.nav_panel("Circos (Density)", ui.output_image("img11", width="100%", height="850px")),
            ),
        ),
        ui.nav_panel(
            "5. Compare Models",
            ui.div("Run the pipeline first. This tab reuses cached model results and does not re-fit any model.", class_="info-box"),
            ui.layout_columns(
                card(
                    ui.div("Comparison Settings", class_="section-title"),
                    ui.input_checkbox_group("cmp_models", "Models to compare", choices=ALL_MODELS, selected=["OLS","MLM","EMMAX","FarmCPU","GEMMA"]),
                    ui.input_radio_buttons("cmp_mode", "Significance definition", choices=["LOD (permutation) — recommended","Bonferroni (0.05/m)","Top-N rank"], selected="LOD (permutation) — recommended"),
                    slider("cmp_top_n", "N / fallback Top-N",5,50,15,1),
                    ui.layout_columns(slider("cmp_w","Width",10,24,16,1),slider("cmp_h","Height",6,16,10,1),slider("cmp_fs","Font Size",6,16,10,1),col_widths=[4,4,4],gap="10px"),
                    ui.input_action_button("cmp_btn", "Find Common MTAs"),
                ),
                card(ui.output_image("cmp_img", width="100%", height="700px")),
                col_widths=[4,8], gap="20px",
            ),
            ui.div("Shared / Unique MTA Table", class_="section-title"),
            ui.output_data_frame("cmp_table"),
            ui.layout_columns(ui.download_button("cmp_fig_dl","Download Common-MTA Figure"), ui.download_button("cmp_table_dl","Download Common-MTA Table"), col_widths=[6,6]),
        ),
    ),
)


def _as_file_obj(file_value):
    """Convert Shiny's input_file() result to the .name interface used by parse_data()."""
    if not file_value:
        return None
    f = file_value[0]
    return SimpleNamespace(name=f["datapath"])


def _num(input, name):
    return input[name]()


def _run_args(input):
    return [
        _as_file_obj(input.file_in()), input.text_in(), input.model_sel(), _num(input,"maf_thr"), _num(input,"top_n_sl"), _num(input,"n_perm"),
        _num(input,"lod_thr_manual"), _num(input,"pve_thr"),
        _num(input,"iter_max_iter"), _num(input,"farmcpu_bin_size"), _num(input,"blink_ld_r2"),
        _num(input,"mrmlm_screen_thresh"), _num(input,"mrmlm_max_candidates"), _num(input,"mrmlm_drop_thresh"),
        _num(input,"mlm_n_pca"), _num(input,"mrmlm_n_pca"), _num(input,"fastmrmlm_n_pca"), _num(input,"gemma_n_pca"),
        _num(input,"mh_w"), _num(input,"mh_h"), _num(input,"mh_fs"), input.mh_fc(), _num(input,"mh_ms"), _num(input,"mh_alpha"), input.mh_bg(), input.mh_sig(), input.mh_sug(), input.mh_dot(), input.mh_panel(), input.mh_show_logp(),
        _num(input,"qq_w"), _num(input,"qq_h"), _num(input,"qq_fs"), input.qq_fc(), _num(input,"qq_ms"), input.qq_dot(), input.qq_line(), input.qq_fill(), input.qq_bg(), input.qq_panel(),
        input.nc_models(), input.nc_theme(), _num(input,"nc_rings_per_panel"), _num(input,"nc_w"), _num(input,"nc_h"), _num(input,"nc_fs"),
        _num(input,"ld_w"), _num(input,"ld_h"), _num(input,"ld_fs"), input.ld_dot(), input.ld_line(),
        _num(input,"pc_w"), _num(input,"pc_h"), _num(input,"pc_fs"),
        _num(input,"ki_w"), _num(input,"ki_h"), _num(input,"ki_fs"), input.ki_dot(), input.ki_bg(), input.ki_panel(),
        _num(input,"ef_w"), _num(input,"ef_h"), _num(input,"ef_fs"), input.ef_pos(), input.ef_neg(), input.ef_fill(), input.ef_bg(),
        _num(input,"co_w"), _num(input,"co_h"), input.co_sig(), input.co_sug(),
        _num(input,"db_w"), _num(input,"db_h"), input.db_dot(),
    ]


def _save_images(imgs, directory):
    paths=[]
    Path(directory).mkdir(parents=True, exist_ok=True)
    for i, im in enumerate(imgs, 1):
        if im is None:
            paths.append(None); continue
        p=Path(directory)/f"plot_{i:02d}.png"
        im.save(p)
        paths.append(str(p))
    return paths


def server(input: Inputs, output: Outputs, session: Session):
    state = reactive.value({})
    output_paths = reactive.value([])
    comparison = reactive.value({})

    @reactive.effect
    @reactive.event(input.run_btn)
    def _run():
        try:
            result = run_gwas(*_run_args(input))
            if len(result) >= 16 and isinstance(result[-1], dict):
                imgs = result[:11]
                paths = _save_images(imgs, tempfile.mkdtemp(prefix="shiny_gwas_plots_"))
                result[ -1 ]["_image_paths"] = paths
                state.set(result[-1])
                output_paths.set(result)
            else:
                state.set({"error": str(result[-1]) if result else "GWAS failed."})
                output_paths.set([])
        except Exception as exc:
            state.set({"error": f"GWAS pipeline error: {exc}"})
            output_paths.set([])

    @render.text
    def summary_out():
        st = state.get()
        if not st:
            return "Ready. Configure the analysis and click Run Complete GWAS Pipeline."
        if "error" in st:
            return st["error"]
        vals = output_paths.get()
        return vals[13] if len(vals) > 13 else "Pipeline completed."

    def image_output(index):
        @render.image
        def _image():
            paths = state.get().get("_image_paths", [])
            if len(paths) >= index and paths[index-1]:
                return {"src": paths[index-1], "width": "100%", "alt": f"GWAS plot {index}"}
            return None
        return _image

    # Register the eleven plot renderers using the output IDs expected by the UI.
    for idx, oid in enumerate(["img1","img2","img3","img4","img5","img6","img7","img8","img9","img10","img11"], 1):
        fn = image_output(idx)
        fn.__name__ = oid
        render.image(fn)

    @reactive.effect
    def _live_manhattan():
        st = state.get()
        if not st or "primary_p" not in st:
            return
        # Read the customization inputs so Shiny invalidates this effect when
        # any Manhattan setting changes; this does NOT call any GWAS model.
        vals = [input.mh_w(),input.mh_h(),input.mh_fs(),input.mh_fc(),input.mh_ms(),input.mh_alpha(),input.mh_bg(),input.mh_sig(),input.mh_sug(),input.mh_dot(),input.mh_panel(),input.mh_show_logp()]
        fig = plot_manhattan(st["primary_p"], st["chr_filt"], st["pos_filt"], "Manhattan Plot", st["primary_model"],
                             build_cfg(vals[0],vals[1],vals[2],vals[3],bg=vals[6],panel=vals[10],sig_col=vals[7],sug_col=vals[8],marker_size=vals[4],alpha=vals[5],dot_col=vals[9]),
                             st["top_n"], lod_threshold=st.get("lod_threshold"), pve_vals=st.get("pve_vals"), pve_threshold=st.get("pve_threshold"), show_logp_axis=bool(vals[11]))
        p=Path(tempfile.mkdtemp(prefix="shiny_manhattan_"))/"manhattan.png"
        fig.savefig(p, dpi=RENDER_DPI, bbox_inches="tight", facecolor=fig.get_facecolor()); plt.close(fig)
        paths=list(st.get("_image_paths", []))
        if paths:
            paths[0]=str(p)
            st["_image_paths"]=paths
            state.set(st)

    @reactive.effect
    @reactive.event(input.cmp_btn)
    def _compare():
        st=state.get()
        if not st or "all_results" not in st:
            comparison.set({"error":"Run the pipeline first."})
            return
        try:
            img, df, fig_path, table_path = run_compare_models(st, input.cmp_models(), input.cmp_mode(), int(input.cmp_top_n()), input.cmp_w(), input.cmp_h(), input.cmp_fs())
            paths=[]
            if img is not None:
                p=Path(tempfile.mkdtemp(prefix="shiny_compare_"))/"common_MTA_figure.png"
                img.save(p); paths.append(str(p))
            comparison.set({"img":paths[0] if paths else None,"df":df,"fig":fig_path,"table":table_path})
        except Exception as exc:
            comparison.set({"error":f"Comparison error: {exc}"})

    @render.image
    def cmp_img():
        c=comparison.get()
        return {"src":c["img"],"width":"100%","alt":"Common MTA comparison"} if c.get("img") else None

    @render.data_frame
    def cmp_table():
        c=comparison.get()
        if c.get("df") is not None:
            return render.DataGrid(c["df"], width="100%", height="500px")
        return render.DataGrid(pd.DataFrame())

    def _download_path(key):
        async def _download():
            vals=output_paths.get()
            path=vals[key] if len(vals)>key else None
            if path and os.path.exists(path):
                with open(path,"rb") as f:
                    yield f.read()
        return _download

    # output_paths indices from run_gwas: 11 Excel, 12 ZIP, 13 TXT, 14 summary.
    for oid, key, filename in [("excel_dl",11,"gwas_results_top15.xlsx"),("zip_dl",12,"gwas_plots.zip"),("summary_dl",13,"gwas_summary.txt")]:
        fn = _download_path(key)
        fn.__name__ = oid
        render.download_button(fn, filename=filename)

    @render.download_button(filename="common_MTA_figure.png")
    async def cmp_fig_dl():
        p=comparison.get().get("fig")
        if p and os.path.exists(p):
            with open(p,"rb") as f: yield f.read()

    @render.download_button(filename="common_MTA_table.xlsx")
    async def cmp_table_dl():
        p=comparison.get().get("table")
        if p and os.path.exists(p):
            with open(p,"rb") as f: yield f.read()


app = App(app_ui, server)
